"""Tests for build_software.py using REAL files - no mocks.

All tests create real Excel files and real HTML files on disk
to verify the actual behavior of the build script.
"""
import pytest
from pathlib import Path
import tempfile
import openpyxl

from build_software import (
    load_software,
    generate_software_item,
    generate_section_content,
    build_software,
)


class TestLoadSoftware:
    """Test loading software data from Excel."""

    @pytest.fixture
    def temp_dir(self):
        """Create a real temporary directory."""
        with tempfile.TemporaryDirectory() as td:
            yield Path(td)

    def test_loads_single_sheet(self, temp_dir):
        """Test loading data from a single sheet."""
        xlsx_path = temp_dir / 'software.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'python'
        ws.append(['name', 'description', 'links_html'])
        ws.append(['TestTool', 'A test tool.', '[<a href="#">Link</a>]'])
        wb.save(xlsx_path)

        data = load_software(xlsx_path)

        assert 'python' in data
        assert len(data['python']) == 1
        assert data['python'][0]['name'] == 'TestTool'

    def test_loads_multiple_sheets(self, temp_dir):
        """Test loading data from multiple sheets."""
        xlsx_path = temp_dir / 'software.xlsx'
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for sheet_name in ['python', 'javascript', 'matlab']:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(['name', 'description', 'links_html'])
            ws.append([f'{sheet_name}_tool', 'Description', ''])

        wb.save(xlsx_path)

        data = load_software(xlsx_path)

        assert len(data) == 3
        assert 'python' in data
        assert 'javascript' in data
        assert 'matlab' in data

    def test_handles_empty_cells(self, temp_dir):
        """Test that empty cells become empty strings."""
        xlsx_path = temp_dir / 'software.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'python'
        ws.append(['name', 'description', 'links_html'])
        ws.append(['Name', None, None])
        wb.save(xlsx_path)

        data = load_software(xlsx_path)

        assert data['python'][0]['description'] == ''
        assert data['python'][0]['links_html'] == ''


class TestGenerateSoftwareItem:
    """Test HTML generation for software items."""

    def test_generates_complete_item(self):
        """Test generating a complete software item."""
        item = {
            'name': 'MyTool',
            'description': 'Does something useful.',
            'links_html': '[<a href="#">GitHub</a>]'
        }

        html = generate_software_item(item)

        assert '<p>' in html
        assert '</p>' in html
        assert '<strong>MyTool.</strong>' in html
        assert 'Does something useful.' in html
        assert '[<a href="#">GitHub</a>]' in html

    def test_handles_missing_links(self):
        """Test item without links."""
        item = {
            'name': 'NoLinks',
            'description': 'A tool without links.',
            'links_html': ''
        }

        html = generate_software_item(item)

        assert '<strong>NoLinks.</strong>' in html
        assert 'A tool without links.' in html
        assert html.count('[') == 0

    def test_preserves_html_in_description(self):
        """Test that HTML in description is preserved."""
        item = {
            'name': 'LinkTool',
            'description': 'Works with <a href="#">other tool</a> and stuff.',
            'links_html': ''
        }

        html = generate_software_item(item)

        assert '<a href="#">other tool</a>' in html


class TestGenerateSectionContent:
    """Test HTML generation for software sections."""

    def test_generates_multiple_items(self):
        """Test generating content with multiple items."""
        items = [
            {'name': 'Tool A', 'description': 'Desc A', 'links_html': ''},
            {'name': 'Tool B', 'description': 'Desc B', 'links_html': ''},
        ]

        html = generate_section_content(items)

        assert html.count('<p>') == 2
        assert 'Tool A' in html
        assert 'Tool B' in html

    def test_handles_empty_list(self):
        """Test generating content with no items."""
        html = generate_section_content([])

        assert html == ''


class TestBuildSoftware:
    """Test full build process."""

    @pytest.fixture
    def temp_dir(self):
        """Create a real temporary directory."""
        with tempfile.TemporaryDirectory() as td:
            yield Path(td)

    def test_builds_complete_page(self, temp_dir):
        """Test building a complete software page."""
        # Create data file
        xlsx_path = temp_dir / 'data.xlsx'
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for sheet_name in ['python', 'javascript', 'matlab']:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(['name', 'description', 'links_html'])
            if sheet_name == 'python':
                ws.append(['TestPython', 'Python tool.', '[<a href="#">Link</a>]'])

        wb.save(xlsx_path)

        # Create template
        template_path = temp_dir / 'template.html'
        template_path.write_text('''<!DOCTYPE html>
<html>
<body>
<div id="python"><!-- PYTHON_CONTENT --></div>
<div id="javascript"><!-- JAVASCRIPT_CONTENT --></div>
<div id="matlab"><!-- MATLAB_CONTENT --></div>
</body>
</html>''')

        # Build
        output_path = temp_dir / 'output.html'
        build_software(xlsx_path, template_path, output_path)

        # Verify
        result = output_path.read_text()
        assert 'TestPython' in result
        assert 'Python tool.' in result

    def test_handles_special_characters(self, temp_dir):
        """Test that special characters are preserved."""
        xlsx_path = temp_dir / 'data.xlsx'
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for sheet_name in ['python', 'javascript', 'matlab']:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(['name', 'description', 'links_html'])
            if sheet_name == 'matlab':
                ws.append(['Sane pColor', "Doesn't look blurry.", ''])

        wb.save(xlsx_path)

        template_path = temp_dir / 'template.html'
        template_path.write_text('''<!DOCTYPE html>
<html>
<body>
<div id="python"><!-- PYTHON_CONTENT --></div>
<div id="javascript"><!-- JAVASCRIPT_CONTENT --></div>
<div id="matlab"><!-- MATLAB_CONTENT --></div>
</body>
</html>''')

        output_path = temp_dir / 'output.html'
        build_software(xlsx_path, template_path, output_path)

        result = output_path.read_text()
        assert "Doesn't" in result


class TestIntegration:
    """Integration tests using actual project files."""

    def test_can_load_real_software_data(self):
        """Test loading the actual software.xlsx file."""
        project_root = Path(__file__).parent.parent
        xlsx_path = project_root / 'data' / 'software.xlsx'

        if not xlsx_path.exists():
            pytest.skip("software.xlsx not found")

        data = load_software(xlsx_path)

        # Should have all expected sections
        assert 'python' in data
        assert 'javascript' in data
        assert 'matlab' in data

        # Should have some content
        assert len(data['python']) > 0
        assert len(data['matlab']) > 0

    def test_can_build_from_real_data(self):
        """Test building from actual project files."""
        project_root = Path(__file__).parent.parent
        data_path = project_root / 'data' / 'software.xlsx'
        template_path = project_root / 'templates' / 'software.html'

        if not data_path.exists() or not template_path.exists():
            pytest.skip("Required files not found")

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / 'software.html'
            build_software(data_path, template_path, output_path)

            # Verify output exists and has content
            assert output_path.exists()
            content = output_path.read_text()
            assert len(content) > 3000  # Should be a substantial page
            assert 'software-list' in content
