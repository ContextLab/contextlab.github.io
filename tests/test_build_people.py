"""Tests for build_people.py using REAL files - no mocks.

All tests create real Excel files and real HTML files on disk
to verify the actual behavior of the build script.
"""

import sys
from pathlib import Path
import tempfile

import pytest
import openpyxl

# Add scripts directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from build_people import (
    load_people,
    generate_director_content,
    generate_member_card,
    generate_members_content,
    generate_alumni_entry,
    generate_alumni_list_content,
    generate_undergrad_entry,
    generate_undergrad_list_content,
    generate_collaborator_entry,
    generate_collaborators_content,
    build_people,
)


class TestLoadPeople:
    """Test loading people data from Excel."""

    @pytest.fixture
    def temp_dir(self):
        """Create a real temporary directory."""
        with tempfile.TemporaryDirectory() as td:
            yield Path(td)

    def test_loads_single_sheet(self, temp_dir):
        """Test loading data from a single sheet."""
        xlsx_path = temp_dir / "people.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "director"
        ws.append(["image", "name", "name_url", "role", "bio", "links_html"])
        ws.append(
            ["jeremy.png", "Jeremy Manning", "", "lab director", "Bio text", "[CV]"]
        )
        wb.save(xlsx_path)

        data = load_people(xlsx_path)

        assert "director" in data
        assert len(data["director"]) == 1
        assert data["director"][0]["name"] == "Jeremy Manning"

    def test_loads_multiple_sheets(self, temp_dir):
        """Test loading data from multiple sheets."""
        xlsx_path = temp_dir / "people.xlsx"
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        assert default_sheet is not None
        wb.remove(default_sheet)

        for sheet_name in ["director", "members", "alumni_postdocs"]:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["name", "role"])
            ws.append([f"{sheet_name} person", "role"])

        wb.save(xlsx_path)

        data = load_people(xlsx_path)

        assert len(data) == 3
        assert "director" in data
        assert "members" in data
        assert "alumni_postdocs" in data

    def test_handles_empty_cells(self, temp_dir):
        """Test that empty cells become empty strings."""
        xlsx_path = temp_dir / "people.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "director"
        ws.append(["image", "name", "name_url", "role", "bio", "links_html"])
        ws.append(["test.png", "Name", None, None, None, None])
        wb.save(xlsx_path)

        data = load_people(xlsx_path)

        assert data["director"][0]["name_url"] == ""
        assert data["director"][0]["role"] == ""


class TestGenerateDirectorContent:
    """Test HTML generation for lab director."""

    def test_generates_director_with_links(self):
        """Test generating director section with links."""
        director = {
            "image": "jeremy.png",
            "name": "jeremy manning",
            "name_url": "",
            "role": "lab director",
            "bio": "Bio paragraph here.",
            "links_html": "CV:http://example.com/cv",
        }

        html = generate_director_content(director)

        assert "two-column lab-director" in html
        assert "images/people/jeremy.png" in html
        assert "jeremy manning" in html
        assert "lab director" in html
        assert "Bio paragraph here." in html
        assert "CV" in html
        assert "http://example.com/cv" in html

    def test_handles_no_links(self):
        """Test director without links."""
        director = {
            "image": "jeremy.png",
            "name": "jeremy manning",
            "name_url": "",
            "role": "lab director",
            "bio": "Bio paragraph.",
            "links_html": "",
        }

        html = generate_director_content(director)

        assert "Bio paragraph." in html
        # Should not have an empty <p></p> for links
        assert html.count("<p>") == 1  # Only the bio paragraph


class TestGenerateMemberCard:
    """Test HTML generation for member cards."""

    def test_generates_basic_card(self):
        """Test generating a basic member card."""
        member = {
            "image": "paxton.png",
            "name": "paxton fitzpatrick",
            "name_url": "http://example.com",
            "role": "grad student",
            "bio": "Paxton is a researcher.",
        }

        html = generate_member_card(member)

        assert "person-card" in html
        assert "images/people/paxton.png" in html
        assert (
            '<a href="http://example.com" target="_blank">paxton fitzpatrick</a>'
            in html
        )
        assert "grad student" in html
        assert "Paxton is a researcher." in html

    def test_handles_no_url(self):
        """Test card without name URL."""
        member = {
            "image": "test.png",
            "name": "test person",
            "name_url": "",
            "role": "undergrad",
            "bio": "Bio here.",
        }

        html = generate_member_card(member)

        assert "<h3>test person | undergrad</h3>" in html
        assert "href=" not in html.split("</h3>")[0].split("<h3>")[1]


class TestGenerateMembersContent:
    """Test HTML generation for members grid."""

    def test_generates_multiple_cards(self):
        """Test generating content with multiple members."""
        members = [
            {
                "image": "a.png",
                "name": "Person A",
                "name_url": "",
                "role": "grad",
                "bio": "A",
            },
            {
                "image": "b.png",
                "name": "Person B",
                "name_url": "",
                "role": "undergrad",
                "bio": "B",
            },
        ]

        html = generate_members_content(members)

        assert html.count("person-card") == 2
        assert "Person A" in html
        assert "Person B" in html

    def test_groups_into_rows_of_three(self):
        """Test that members are grouped into grids of 3."""
        members = [
            {
                "image": f"{i}.png",
                "name": f"Person {i}",
                "name_url": "",
                "role": "role",
                "bio": "Bio",
            }
            for i in range(7)
        ]

        html = generate_members_content(members)

        # 7 members = 3 grids (3, 3, 1)
        assert html.count("people-grid") == 3

    def test_handles_empty_list(self):
        """Test generating content with no members."""
        html = generate_members_content([])

        assert html == ""


class TestGenerateAlumniEntry:
    """Test HTML generation for alumni entries."""

    def test_generates_entry_with_all_fields(self):
        """Test alumni entry with name URL and position URL."""
        alum = {
            "name": "Andrew Heusser",
            "name_url": "http://example.com",
            "years": "2016-2018",
            "current_position": "Akili",
            "current_position_url": "http://akili.com",
        }

        html = generate_alumni_entry(alum)

        assert '<a href="http://example.com" target="_blank">Andrew Heusser</a>' in html
        assert "2016-2018" in html
        assert 'now at <a href="http://akili.com" target="_blank">Akili</a>' in html

    def test_handles_no_urls(self):
        """Test alumni entry without URLs."""
        alum = {
            "name": "Mark Taylor",
            "name_url": "",
            "years": "QBS Masters 2021",
            "current_position": "",
            "current_position_url": "",
        }

        html = generate_alumni_entry(alum)

        assert html == "Mark Taylor (QBS Masters 2021)"

    def test_handles_position_without_prefix(self):
        """Test alumni with position that should get 'now at' prefix."""
        alum = {
            "name": "Kirsten Ziman",
            "name_url": "http://example.com",
            "years": "2016-2017",
            "current_position": "Princeton University",
            "current_position_url": "http://princeton.edu",
        }

        html = generate_alumni_entry(alum)

        assert (
            'now at <a href="http://princeton.edu" target="_blank">Princeton University</a>'
            in html
        )


class TestGenerateUndergradEntry:
    """Test HTML generation for undergraduate alumni."""

    def test_generates_entry_with_years(self):
        """Test undergrad entry with years."""
        alum = {"name": "Jane Doe", "years": "2020-2021"}

        html = generate_undergrad_entry(alum)

        assert html == "Jane Doe (2020-2021)"

    def test_handles_no_years(self):
        """Test undergrad entry without years."""
        alum = {"name": "Jane Doe", "years": ""}

        html = generate_undergrad_entry(alum)

        assert html == "Jane Doe"


class TestGenerateCollaboratorEntry:
    """Test HTML generation for collaborator entries."""

    def test_generates_entry_with_url(self):
        """Test collaborator entry with URL."""
        collab = {
            "name": "Memory Lab",
            "url": "http://memory.example.com",
            "description": "Memory Lab, University (Director: Someone)",
        }

        html = generate_collaborator_entry(collab)

        assert "<p>" in html
        assert "</p>" in html
        assert (
            '<a href="http://memory.example.com" target="_blank">Memory Lab</a>' in html
        )

    def test_handles_no_url(self):
        """Test collaborator entry without URL."""
        collab = {"name": "Test Lab", "url": "", "description": "Test Lab description"}

        html = generate_collaborator_entry(collab)

        assert "<p>Test Lab description</p>" == html


class TestBuildPeople:
    """Test full build process."""

    @pytest.fixture
    def temp_dir(self):
        """Create a real temporary directory."""
        with tempfile.TemporaryDirectory() as td:
            yield Path(td)

    def test_builds_complete_page(self, temp_dir):
        """Test building a complete people page."""
        # Create data file
        xlsx_path = temp_dir / "data.xlsx"
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        assert default_sheet is not None
        wb.remove(default_sheet)

        # Create all required sheets
        for sheet_name in [
            "director",
            "members",
            "alumni_postdocs",
            "alumni_grads",
            "alumni_managers",
            "alumni_undergrads",
            "collaborators",
        ]:
            ws = wb.create_sheet(title=sheet_name)
            if sheet_name in ["director", "members"]:
                ws.append(["image", "name", "name_url", "role", "bio", "links_html"])
                if sheet_name == "director":
                    ws.append(["dir.png", "Director Name", "", "director", "Bio", ""])
                else:
                    ws.append(["member.png", "Member Name", "", "role", "Bio", ""])
            elif sheet_name == "alumni_undergrads":
                ws.append(["name", "years"])
                ws.append(["Undergrad Name", "2020"])
            elif sheet_name == "collaborators":
                ws.append(["name", "url", "description"])
                ws.append(["Lab Name", "http://example.com", "Lab Name, Description"])
            else:
                ws.append(
                    [
                        "name",
                        "name_url",
                        "years",
                        "current_position",
                        "current_position_url",
                    ]
                )
                ws.append(["Alumni Name", "", "2020", "", ""])

        wb.save(xlsx_path)

        # Create template
        template_path = temp_dir / "template.html"
        template_path.write_text("""<!DOCTYPE html>
<html>
<body>
<section id="lab-members"><!-- DIRECTOR_CONTENT --></section>
<section><!-- MEMBERS_CONTENT --></section>
<div><p><!-- ALUMNI_POSTDOCS_CONTENT --></p></div>
<div><p><!-- ALUMNI_GRADS_CONTENT --></p></div>
<div><p><!-- ALUMNI_MANAGERS_CONTENT --></p></div>
<div><p><!-- ALUMNI_UNDERGRADS_CONTENT --></p></div>
<div><!-- COLLABORATORS_CONTENT --></div>
</body>
</html>""")

        # Build
        output_path = temp_dir / "output.html"
        build_people(xlsx_path, template_path, output_path)

        # Verify
        result = output_path.read_text()
        assert "Director Name" in result
        assert "Member Name" in result
        assert "Undergrad Name" in result
        assert "Lab Name" in result


class TestIntegration:
    """Integration tests using actual project files."""

    def test_can_load_real_people_data(self):
        """Test loading the actual people.xlsx file."""
        project_root = Path(__file__).parent.parent
        xlsx_path = project_root / "data" / "people.xlsx"

        if not xlsx_path.exists():
            pytest.skip("people.xlsx not found")

        data = load_people(xlsx_path)

        # Should have all expected sections
        assert "director" in data
        assert "members" in data
        assert "alumni_postdocs" in data
        assert "alumni_grads" in data
        assert "alumni_managers" in data
        assert "alumni_undergrads" in data
        assert "collaborators" in data

        # Should have some content
        assert len(data["director"]) == 1
        assert len(data["members"]) > 0

    def test_can_build_from_real_data(self):
        """Test building from actual project files."""
        project_root = Path(__file__).parent.parent
        data_path = project_root / "data" / "people.xlsx"
        template_path = project_root / "templates" / "people.html"

        if not data_path.exists() or not template_path.exists():
            pytest.skip("Required files not found")

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "people.html"
            build_people(data_path, template_path, output_path)

            # Verify output exists and has content
            assert output_path.exists()
            content = output_path.read_text()
            assert len(content) > 5000  # Should be a substantial page
            assert "person-card" in content
            assert "lab-director" in content


class TestEntryPointConsistency:
    """Both build.py and build_people.py must produce the same people.html.

    Regression tests for the ordering thrash described in issue #16: build.py
    used to call build_people() without a cv_path, which dropped the CV-derived
    ordering of undergrad alumni and made the two entry points fight over the
    file, producing a 152-line diff every time the other one ran.
    """

    def test_committed_people_html_is_the_cv_ordered_build(self):
        """The committed people.html must match a build that uses the CV order."""
        project_root = Path(__file__).parent.parent
        data_path = project_root / "data" / "people.xlsx"
        template_path = project_root / "templates" / "people.html"
        cv_path = project_root / "documents" / "JRM_CV.tex"
        committed = project_root / "people.html"

        for required in (data_path, template_path, cv_path, committed):
            if not required.exists():
                pytest.skip(f"{required.name} not found")

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "people.html"
            build_people(data_path, template_path, output_path, cv_path)

            assert output_path.read_text(encoding="utf-8") == committed.read_text(
                encoding="utf-8"
            ), "committed people.html does not match the CV-ordered build"

    def test_cv_ordering_actually_changes_the_output(self):
        """Guard the guard: if these matched, the test above would prove nothing."""
        project_root = Path(__file__).parent.parent
        data_path = project_root / "data" / "people.xlsx"
        template_path = project_root / "templates" / "people.html"
        cv_path = project_root / "documents" / "JRM_CV.tex"

        for required in (data_path, template_path, cv_path):
            if not required.exists():
                pytest.skip(f"{required.name} not found")

        with tempfile.TemporaryDirectory() as td:
            with_cv = Path(td) / "with_cv.html"
            without_cv = Path(td) / "without_cv.html"
            build_people(data_path, template_path, with_cv, cv_path)
            build_people(data_path, template_path, without_cv)

            assert with_cv.read_text(encoding="utf-8") != without_cv.read_text(
                encoding="utf-8"
            ), "cv_path no longer affects ordering; this test needs rethinking"

    def test_build_py_leaves_people_html_unchanged(self):
        """Running build.py must not rewrite the committed people.html."""
        import subprocess

        project_root = Path(__file__).parent.parent
        scripts_dir = project_root / "scripts"
        committed = project_root / "people.html"

        if not committed.exists():
            pytest.skip("people.html not found")

        before = committed.read_bytes()
        result = subprocess.run(
            [sys.executable, "build.py"],
            cwd=scripts_dir,
            capture_output=True,
            text=True,
        )
        after = committed.read_bytes()

        # Restore before asserting so a failure can't leave the tree dirty
        if after != before:
            committed.write_bytes(before)

        assert result.returncode == 0, f"build.py failed: {result.stderr}"
        assert after == before, "build.py rewrote people.html (ordering thrash is back)"

    def test_build_people_py_leaves_people_html_unchanged(self):
        """The symmetric check, and the one that has teeth in CI.

        build-content.yml runs build.py *before* pytest, so a test comparing
        build.py's output to the file on disk is tautological there. Running
        the OTHER entry point afterwards is not: it fails if build_people.py
        would write something different from what build.py just wrote.
        """
        import subprocess

        project_root = Path(__file__).parent.parent
        scripts_dir = project_root / "scripts"
        committed = project_root / "people.html"

        if not committed.exists():
            pytest.skip("people.html not found")

        before = committed.read_bytes()
        result = subprocess.run(
            [sys.executable, "build_people.py"],
            cwd=scripts_dir,
            capture_output=True,
            text=True,
        )
        after = committed.read_bytes()

        if after != before:
            committed.write_bytes(before)

        assert result.returncode == 0, f"build_people.py failed: {result.stderr}"
        assert after == before, "build_people.py disagrees with build.py"


class TestCvOrderingRobustness:
    """Ordering must not depend on capitalization agreeing across sources."""

    def test_cv_position_lookup_ignores_case(self):
        """data/people.xlsx spells one alumnus differently from the CV."""
        project_root = Path(__file__).parent.parent
        data_path = project_root / "data" / "people.xlsx"
        cv_path = project_root / "documents" / "JRM_CV.tex"

        for required in (data_path, cv_path):
            if not required.exists():
                pytest.skip(f"{required.name} not found")

        from build_people import parse_cv_undergrad_order

        cv_names = {n.casefold() for n in parse_cv_undergrad_order(cv_path)}
        sheet_names = [p["name"] for p in load_people(data_path)["alumni_undergrads"]]

        unmatched = [n for n in sheet_names if n.casefold() not in cv_names]
        assert not unmatched, f"undergrad alumni missing from the CV: {unmatched}"

    def test_case_mismatch_still_lands_in_cv_position(self):
        """A name that differs only in case must not sort to the end."""
        from build_people import generate_undergrad_list_content

        alumni = [
            {"name": "Zoe Zebra", "years": "2020 - 2021"},
            {"name": "Evan Mcdermid", "years": "2020 - 2021"},
        ]
        html = generate_undergrad_list_content(alumni, ["Evan McDermid", "Zoe Zebra"])

        assert html.index("Evan Mcdermid") < html.index("Zoe Zebra")
