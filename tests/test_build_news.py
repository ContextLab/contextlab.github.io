"""Tests for build_news.py using REAL files - no mocks.

All tests create real Excel files and real HTML files on disk
to verify the actual behavior of the build script.
"""
import pytest
from pathlib import Path
import tempfile
import openpyxl

from build_news import (
    load_news,
    generate_news_item,
    generate_news_content,
    build_news,
)


class TestLoadNews:
    """Test loading news from Excel."""

    @pytest.fixture
    def temp_dir(self):
        """Create a real temporary directory."""
        with tempfile.TemporaryDirectory() as td:
            yield Path(td)

    def test_loads_news_items(self, temp_dir):
        """Test loading data from a single sheet."""
        xlsx_path = temp_dir / 'news.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['date', 'title', 'title_url', 'content', 'links_html', 'image'])
        ws.append(['2024-01-15', 'Test News', 'http://example.com', 'Content here', '', 'test.png'])
        wb.save(xlsx_path)

        items = load_news(xlsx_path)

        assert len(items) == 1
        assert items[0]['title'] == 'Test News'
        assert items[0]['date'] == '2024-01-15'

    def test_sorts_by_date_descending(self, temp_dir):
        """Test that news items are sorted by date (most recent first)."""
        xlsx_path = temp_dir / 'news.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['date', 'title', 'title_url', 'content', 'links_html', 'image'])
        ws.append(['2020-01-01', 'Old News', '', 'Old content', '', ''])
        ws.append(['2024-06-15', 'Recent News', '', 'Recent content', '', ''])
        ws.append(['2022-05-10', 'Middle News', '', 'Middle content', '', ''])
        wb.save(xlsx_path)

        items = load_news(xlsx_path)

        assert len(items) == 3
        assert items[0]['title'] == 'Recent News'
        assert items[1]['title'] == 'Middle News'
        assert items[2]['title'] == 'Old News'

    def test_handles_empty_cells(self, temp_dir):
        """Test that empty cells become empty strings."""
        xlsx_path = temp_dir / 'news.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['date', 'title', 'title_url', 'content', 'links_html', 'image'])
        ws.append(['2024-01-15', 'Title', None, 'Content', None, None])
        wb.save(xlsx_path)

        items = load_news(xlsx_path)

        assert items[0]['title_url'] == ''
        assert items[0]['links_html'] == ''
        assert items[0]['image'] == ''

    def test_handles_missing_dates(self, temp_dir):
        """Test that items without dates sort to the end."""
        xlsx_path = temp_dir / 'news.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['date', 'title', 'title_url', 'content', 'links_html', 'image'])
        ws.append(['', 'No Date News', '', 'Content', '', ''])
        ws.append(['2024-01-15', 'Dated News', '', 'Content', '', ''])
        wb.save(xlsx_path)

        items = load_news(xlsx_path)

        assert len(items) == 2
        assert items[0]['title'] == 'Dated News'
        assert items[1]['title'] == 'No Date News'


class TestGenerateNewsItem:
    """Test HTML generation for news items."""

    def test_generates_basic_item(self):
        """Test generating a basic news item."""
        item = {
            'image': 'test.png',
            'title': 'Test News',
            'title_url': '',
            'content': 'This is the content.',
            'links_html': ''
        }

        html = generate_news_item(item)

        assert '<div class="news-item">' in html
        assert 'images/news/test.png' in html
        assert 'Test News' in html
        assert 'This is the content.' in html

    def test_generates_linked_title(self):
        """Test generating news item with linked title."""
        item = {
            'image': 'test.png',
            'title': 'Linked News',
            'title_url': 'http://example.com',
            'content': 'Content here.',
            'links_html': ''
        }

        html = generate_news_item(item)

        assert '<a href="http://example.com" target="_blank">Linked News</a>' in html

    def test_generates_with_links_html(self):
        """Test generating news item with extra links."""
        item = {
            'image': 'test.png',
            'title': 'News with Links',
            'title_url': '',
            'content': 'Content here.',
            'links_html': '[<a href="http://example.com">READ MORE</a>]'
        }

        html = generate_news_item(item)

        # Should have the links in a separate paragraph
        assert 'READ MORE' in html
        assert html.count('<p>') == 2

    def test_handles_no_links_html(self):
        """Test that no extra paragraph when links_html is empty."""
        item = {
            'image': 'test.png',
            'title': 'News',
            'title_url': '',
            'content': 'Content.',
            'links_html': ''
        }

        html = generate_news_item(item)

        # Should only have content paragraph, not links paragraph
        assert html.count('<p>') == 1


class TestGenerateNewsContent:
    """Test HTML generation for multiple news items."""

    def test_generates_multiple_items(self):
        """Test generating content with multiple news items."""
        items = [
            {'image': 'a.png', 'title': 'News A', 'title_url': '', 'content': 'Content A', 'links_html': ''},
            {'image': 'b.png', 'title': 'News B', 'title_url': '', 'content': 'Content B', 'links_html': ''},
        ]

        html = generate_news_content(items)

        assert html.count('news-item') == 2
        assert 'News A' in html
        assert 'News B' in html

    def test_handles_empty_list(self):
        """Test generating content with no news items."""
        html = generate_news_content([])

        assert html == ''


class TestBuildNews:
    """Test full build process."""

    @pytest.fixture
    def temp_dir(self):
        """Create a real temporary directory."""
        with tempfile.TemporaryDirectory() as td:
            yield Path(td)

    def test_builds_complete_page(self, temp_dir):
        """Test building a complete news page."""
        # Create data file
        xlsx_path = temp_dir / 'data.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['date', 'title', 'title_url', 'content', 'links_html', 'image'])
        ws.append(['2024-06-15', 'Test News', 'http://example.com', 'News content here.', '', 'news.png'])
        wb.save(xlsx_path)

        # Create template
        template_path = temp_dir / 'template.html'
        template_path.write_text('''<!DOCTYPE html>
<html>
<body>
<div class="news-list">
<!-- NEWS_CONTENT -->
</div>
</body>
</html>''')

        # Build
        output_path = temp_dir / 'output.html'
        build_news(xlsx_path, template_path, output_path)

        # Verify
        result = output_path.read_text()
        assert 'Test News' in result
        assert 'news.png' in result
        assert 'News content here.' in result

    def test_handles_special_characters(self, temp_dir):
        """Test that special characters are preserved."""
        xlsx_path = temp_dir / 'data.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['date', 'title', 'title_url', 'content', 'links_html', 'image'])
        ws.append(['2024-01-01', "News & O'Brien's \"quoted\"", '', "Author & O'Brien", '', 'test.png'])
        wb.save(xlsx_path)

        template_path = temp_dir / 'template.html'
        template_path.write_text('''<!DOCTYPE html>
<html>
<body>
<div class="news-list">
<!-- NEWS_CONTENT -->
</div>
</body>
</html>''')

        output_path = temp_dir / 'output.html'
        build_news(xlsx_path, template_path, output_path)

        result = output_path.read_text()
        assert '&' in result
        assert "'" in result
        assert '"' in result


class TestIntegration:
    """Integration tests using actual project files."""

    def test_can_load_real_news_data(self):
        """Test loading the actual news.xlsx file."""
        project_root = Path(__file__).parent.parent
        xlsx_path = project_root / 'data' / 'news.xlsx'

        if not xlsx_path.exists():
            pytest.skip("news.xlsx not found")

        items = load_news(xlsx_path)

        # Should have news items
        assert len(items) > 0

        # First item should have required fields
        assert items[0].get('title')
        assert items[0].get('content')

    def test_news_items_sorted_correctly(self):
        """Test that real news items are sorted by date."""
        project_root = Path(__file__).parent.parent
        xlsx_path = project_root / 'data' / 'news.xlsx'

        if not xlsx_path.exists():
            pytest.skip("news.xlsx not found")

        items = load_news(xlsx_path)

        # Verify descending date order
        from datetime import datetime
        dates = []
        for item in items:
            date_str = item.get('date', '')
            if date_str:
                try:
                    dates.append(datetime.strptime(str(date_str), '%Y-%m-%d'))
                except ValueError:
                    pass

        # Check dates are in descending order
        for i in range(len(dates) - 1):
            assert dates[i] >= dates[i + 1], f"Dates not in descending order at index {i}"

    def test_can_build_from_real_data(self):
        """Test building from actual project files."""
        project_root = Path(__file__).parent.parent
        data_path = project_root / 'data' / 'news.xlsx'
        template_path = project_root / 'templates' / 'news.html'

        if not data_path.exists() or not template_path.exists():
            pytest.skip("Required files not found")

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / 'news.html'
            build_news(data_path, template_path, output_path)

            # Verify output exists and has content
            assert output_path.exists()
            content = output_path.read_text()
            assert len(content) > 1000  # Should be a substantial page
            assert 'news-item' in content

    def test_generated_html_structure(self):
        """Test that generated HTML has expected structure."""
        project_root = Path(__file__).parent.parent
        data_path = project_root / 'data' / 'news.xlsx'
        template_path = project_root / 'templates' / 'news.html'

        if not data_path.exists() or not template_path.exists():
            pytest.skip("Required files not found")

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / 'news.html'
            build_news(data_path, template_path, output_path)

            content = output_path.read_text()

            # Should have Bluesky section
            assert 'bluesky-feed' in content
            assert 'bluesky-section' in content

            # Should have news items
            assert 'news-item' in content
            assert 'news-thumbnail' in content
            assert 'news-content' in content
