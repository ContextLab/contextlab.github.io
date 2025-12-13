"""Tests for build_publications.py using REAL files - no mocks.

All tests create real Excel files and real HTML files on disk
to verify the actual behavior of the build script.
"""
import pytest
from pathlib import Path
import tempfile
import openpyxl

from build_publications import (
    load_publications,
    generate_publication_card,
    generate_section_content,
    build_publications,
)


class TestLoadPublications:
    """Test loading publications from Excel."""

    @pytest.fixture
    def temp_dir(self):
        """Create a real temporary directory."""
        with tempfile.TemporaryDirectory() as td:
            yield Path(td)

    def test_loads_single_sheet(self, temp_dir):
        """Test loading data from a single sheet."""
        xlsx_path = temp_dir / 'pubs.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'papers'
        ws.append(['image', 'title', 'title_url', 'citation', 'links_html'])
        ws.append(['test.png', 'Test Paper', 'http://example.com', 'Author (2024)', '[PDF]'])
        wb.save(xlsx_path)

        data = load_publications(xlsx_path)

        assert 'papers' in data
        assert len(data['papers']) == 1
        assert data['papers'][0]['title'] == 'Test Paper'

    def test_loads_multiple_sheets(self, temp_dir):
        """Test loading data from multiple sheets."""
        xlsx_path = temp_dir / 'pubs.xlsx'
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create multiple sheets
        for sheet_name in ['papers', 'chapters', 'talks']:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(['image', 'title', 'title_url', 'citation', 'links_html'])
            ws.append([f'{sheet_name}.png', f'{sheet_name} title', '', '', ''])

        wb.save(xlsx_path)

        data = load_publications(xlsx_path)

        assert len(data) == 3
        assert 'papers' in data
        assert 'chapters' in data
        assert 'talks' in data

    def test_handles_empty_cells(self, temp_dir):
        """Test that empty cells become empty strings."""
        xlsx_path = temp_dir / 'pubs.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'papers'
        ws.append(['image', 'title', 'title_url', 'citation', 'links_html'])
        ws.append(['test.png', 'Title', None, None, None])
        wb.save(xlsx_path)

        data = load_publications(xlsx_path)

        assert data['papers'][0]['title_url'] == ''
        assert data['papers'][0]['citation'] == ''
        assert data['papers'][0]['links_html'] == ''


class TestGeneratePublicationCard:
    """Test HTML generation for publication cards."""

    def test_generates_basic_card(self):
        """Test generating a basic publication card."""
        pub = {
            'image': 'test.png',
            'title': 'Test Paper',
            'title_url': 'http://example.com',
            'citation': 'Author (2024) Test Paper',
            'links_html': '[<a href="#">PDF</a>]'
        }

        html = generate_publication_card(pub)

        assert '<div class="publication-card">' in html
        assert 'images/publications/test.png' in html
        assert 'Test Paper' in html
        assert 'http://example.com' in html
        assert 'Author (2024) Test Paper' in html
        assert '[<a href="#">PDF</a>]' in html

    def test_handles_no_url(self):
        """Test card with no title URL."""
        pub = {
            'image': 'test.png',
            'title': 'Test Paper',
            'title_url': '',
            'citation': 'Citation',
            'links_html': ''
        }

        html = generate_publication_card(pub)

        assert '<h4>Test Paper</h4>' in html
        assert 'href=' not in html.split('<h4>')[1].split('</h4>')[0]

    def test_handles_no_links(self):
        """Test card with no publication links."""
        pub = {
            'image': 'test.png',
            'title': 'Test Paper',
            'title_url': '',
            'citation': 'Citation',
            'links_html': ''
        }

        html = generate_publication_card(pub)

        assert 'publication-links' not in html

    def test_preserves_html_in_citation(self):
        """Test that HTML in citation is preserved."""
        pub = {
            'image': 'test.png',
            'title': 'Test',
            'title_url': '',
            'citation': 'Author (2024) <em>Journal Name</em>, 1(1): 1-10.',
            'links_html': ''
        }

        html = generate_publication_card(pub)

        assert '<em>Journal Name</em>' in html


class TestGenerateSectionContent:
    """Test HTML generation for publication sections."""

    def test_generates_multiple_cards(self):
        """Test generating content with multiple publications."""
        pubs = [
            {'image': 'a.png', 'title': 'Paper A', 'title_url': '', 'citation': 'A', 'links_html': ''},
            {'image': 'b.png', 'title': 'Paper B', 'title_url': '', 'citation': 'B', 'links_html': ''},
        ]

        html = generate_section_content(pubs)

        assert html.count('publication-card') == 2
        assert 'Paper A' in html
        assert 'Paper B' in html

    def test_handles_empty_list(self):
        """Test generating content with no publications."""
        html = generate_section_content([])

        assert html == ''


class TestBuildPublications:
    """Test full build process."""

    @pytest.fixture
    def temp_dir(self):
        """Create a real temporary directory."""
        with tempfile.TemporaryDirectory() as td:
            yield Path(td)

    def test_builds_complete_page(self, temp_dir):
        """Test building a complete publications page."""
        # Create data file
        xlsx_path = temp_dir / 'data.xlsx'
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for sheet_name in ['papers', 'chapters', 'dissertations', 'talks', 'courses', 'posters']:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(['image', 'title', 'title_url', 'citation', 'links_html'])
            if sheet_name == 'papers':
                ws.append(['paper.png', 'Test Paper', 'http://example.com',
                          'Author (2024) Test Paper. <em>Journal</em>.',
                          '[<a href="#">PDF</a>]'])

        wb.save(xlsx_path)

        # Create template
        template_path = temp_dir / 'template.html'
        template_path.write_text('''<!DOCTYPE html>
<html>
<body>
<div id="papers"><!-- PAPERS_CONTENT --></div>
<div id="chapters"><!-- CHAPTERS_CONTENT --></div>
<div id="dissertations"><!-- DISSERTATIONS_CONTENT --></div>
<div id="talks"><!-- TALKS_CONTENT --></div>
<div id="courses"><!-- COURSES_CONTENT --></div>
<div id="posters"><!-- POSTERS_CONTENT --></div>
</body>
</html>''')

        # Build
        output_path = temp_dir / 'output.html'
        build_publications(xlsx_path, template_path, output_path)

        # Verify
        result = output_path.read_text()
        assert 'Test Paper' in result
        assert 'paper.png' in result
        assert '<em>Journal</em>' in result

    def test_handles_special_characters(self, temp_dir):
        """Test that special characters are preserved."""
        xlsx_path = temp_dir / 'data.xlsx'
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for sheet_name in ['papers', 'chapters', 'dissertations', 'talks', 'courses', 'posters']:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(['image', 'title', 'title_url', 'citation', 'links_html'])
            if sheet_name == 'papers':
                ws.append(['test.png', "Paper & O'Brien's \"quoted\"", '',
                          "Author & O'Brien (2024)", ''])

        wb.save(xlsx_path)

        template_path = temp_dir / 'template.html'
        template_path.write_text('''<!DOCTYPE html>
<html>
<body>
<div id="papers"><!-- PAPERS_CONTENT --></div>
<div id="chapters"><!-- CHAPTERS_CONTENT --></div>
<div id="dissertations"><!-- DISSERTATIONS_CONTENT --></div>
<div id="talks"><!-- TALKS_CONTENT --></div>
<div id="courses"><!-- COURSES_CONTENT --></div>
<div id="posters"><!-- POSTERS_CONTENT --></div>
</body>
</html>''')

        output_path = temp_dir / 'output.html'
        build_publications(xlsx_path, template_path, output_path)

        result = output_path.read_text()
        assert '&' in result
        assert "'" in result
        assert '"' in result


class TestIntegration:
    """Integration tests using actual project files."""

    def test_can_load_real_publications_data(self):
        """Test loading the actual publications.xlsx file."""
        project_root = Path(__file__).parent.parent
        xlsx_path = project_root / 'data' / 'publications.xlsx'

        if not xlsx_path.exists():
            pytest.skip("publications.xlsx not found")

        data = load_publications(xlsx_path)

        # Should have all expected sections
        assert 'papers' in data
        assert 'chapters' in data
        assert 'dissertations' in data
        assert 'talks' in data
        assert 'courses' in data
        assert 'posters' in data

        # Should have some content
        assert len(data['papers']) > 0

    def test_can_build_from_real_data(self):
        """Test building from actual project files."""
        project_root = Path(__file__).parent.parent
        data_path = project_root / 'data' / 'publications.xlsx'
        template_path = project_root / 'templates' / 'publications.html'

        if not data_path.exists() or not template_path.exists():
            pytest.skip("Required files not found")

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / 'publications.html'
            build_publications(data_path, template_path, output_path)

            # Verify output exists and has content
            assert output_path.exists()
            content = output_path.read_text()
            assert len(content) > 10000  # Should be a substantial page
            assert 'publication-card' in content
