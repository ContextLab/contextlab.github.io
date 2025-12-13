"""Tests for utils.py using REAL files - no mocks.

All tests create real Excel files and real HTML files on disk
to verify the actual behavior of the utility functions.
"""
import pytest
from pathlib import Path
import tempfile
import openpyxl

from utils import (
    load_spreadsheet,
    inject_content,
    validate_required_fields,
    validate_url_format,
    check_file_exists,
)


class TestLoadSpreadsheet:
    """Test spreadsheet loading with real Excel files."""

    @pytest.fixture
    def temp_dir(self):
        """Create a real temporary directory."""
        with tempfile.TemporaryDirectory() as td:
            yield Path(td)

    def test_loads_simple_spreadsheet(self, temp_dir):
        """Test loading a simple spreadsheet with basic data."""
        filepath = temp_dir / 'test.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['name', 'age', 'city'])
        ws.append(['Alice', 30, 'Boston'])
        ws.append(['Bob', 25, 'Chicago'])
        wb.save(filepath)

        rows = load_spreadsheet(filepath)

        assert len(rows) == 2
        assert rows[0] == {'name': 'Alice', 'age': 30, 'city': 'Boston'}
        assert rows[1] == {'name': 'Bob', 'age': 25, 'city': 'Chicago'}

    def test_handles_empty_cells(self, temp_dir):
        """Test that empty cells become empty strings."""
        filepath = temp_dir / 'test.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['title', 'optional', 'required'])
        ws.append(['Paper A', None, 'yes'])
        ws.append(['Paper B', 'has value', 'yes'])
        wb.save(filepath)

        rows = load_spreadsheet(filepath)

        assert rows[0]['optional'] == ''
        assert rows[1]['optional'] == 'has value'

    def test_skips_empty_rows(self, temp_dir):
        """Test that completely empty rows are skipped."""
        filepath = temp_dir / 'test.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['col1', 'col2'])
        ws.append(['data1', 'data2'])
        ws.append([None, None])  # Empty row
        ws.append(['data3', 'data4'])
        wb.save(filepath)

        rows = load_spreadsheet(filepath)

        assert len(rows) == 2
        assert rows[0]['col1'] == 'data1'
        assert rows[1]['col1'] == 'data3'

    def test_preserves_html_in_cells(self, temp_dir):
        """Test that HTML content in cells is preserved."""
        filepath = temp_dir / 'test.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['title', 'description'])
        ws.append([
            'Paper with <em>emphasis</em>',
            '<a href="http://example.com">Link</a> text'
        ])
        wb.save(filepath)

        rows = load_spreadsheet(filepath)

        assert '<em>' in rows[0]['title']
        assert '<a href=' in rows[0]['description']

    def test_handles_numeric_values(self, temp_dir):
        """Test that numeric values are preserved as numbers."""
        filepath = temp_dir / 'test.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['name', 'year', 'score'])
        ws.append(['Test', 2024, 95.5])
        wb.save(filepath)

        rows = load_spreadsheet(filepath)

        assert rows[0]['year'] == 2024
        assert rows[0]['score'] == 95.5
        assert isinstance(rows[0]['year'], int)
        assert isinstance(rows[0]['score'], float)

    def test_raises_on_missing_file(self, temp_dir):
        """Test that FileNotFoundError is raised for missing files."""
        with pytest.raises(FileNotFoundError):
            load_spreadsheet(temp_dir / 'nonexistent.xlsx')

    def test_raises_on_empty_header(self, temp_dir):
        """Test that empty header cells raise an error."""
        filepath = temp_dir / 'test.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['col1', None, 'col3'])  # Empty header
        ws.append(['a', 'b', 'c'])
        wb.save(filepath)

        with pytest.raises(ValueError, match="empty header"):
            load_spreadsheet(filepath)

    def test_handles_special_characters(self, temp_dir):
        """Test handling of special characters in data."""
        filepath = temp_dir / 'test.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['title', 'authors'])
        ws.append(['Paper & Analysis', "O'Brien, Smith"])
        ws.append(['Title "quoted"', 'Author <special>'])
        wb.save(filepath)

        rows = load_spreadsheet(filepath)

        assert '&' in rows[0]['title']
        assert "'" in rows[0]['authors']
        assert '"' in rows[1]['title']
        assert '<' in rows[1]['authors']


class TestInjectContent:
    """Test content injection into HTML templates."""

    @pytest.fixture
    def temp_dir(self):
        """Create a real temporary directory."""
        with tempfile.TemporaryDirectory() as td:
            yield Path(td)

    def test_injects_single_marker(self, temp_dir):
        """Test injecting content at a single marker."""
        template = temp_dir / 'template.html'
        template.write_text('<div><!-- CONTENT --></div>')

        output = temp_dir / 'output.html'
        inject_content(template, output, {'CONTENT': '<p>Hello</p>'})

        assert output.read_text() == '<div><p>Hello</p></div>'

    def test_injects_multiple_markers(self, temp_dir):
        """Test injecting at multiple markers."""
        template = temp_dir / 'template.html'
        template.write_text('<!-- A -->|<!-- B -->|<!-- C -->')

        output = temp_dir / 'output.html'
        inject_content(template, output, {
            'A': 'First',
            'B': 'Second',
            'C': 'Third'
        })

        assert output.read_text() == 'First|Second|Third'

    def test_preserves_surrounding_content(self, temp_dir):
        """Test that content around markers is preserved."""
        template = temp_dir / 'template.html'
        template.write_text('''<!DOCTYPE html>
<html>
<head><title>Test</title></head>
<body>
<header>Header Content</header>
<!-- MAIN -->
<footer>Footer Content</footer>
</body>
</html>''')

        output = temp_dir / 'output.html'
        inject_content(template, output, {'MAIN': '<main>Injected</main>'})

        result = output.read_text()
        assert '<header>Header Content</header>' in result
        assert '<main>Injected</main>' in result
        assert '<footer>Footer Content</footer>' in result

    def test_raises_on_missing_marker(self, temp_dir):
        """Test that missing marker raises ValueError."""
        template = temp_dir / 'template.html'
        template.write_text('<div>No markers here</div>')

        output = temp_dir / 'output.html'
        with pytest.raises(ValueError, match="not found"):
            inject_content(template, output, {'MISSING': 'content'})

    def test_raises_on_missing_template(self, temp_dir):
        """Test that missing template raises FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            inject_content(
                temp_dir / 'nonexistent.html',
                temp_dir / 'output.html',
                {'X': 'y'}
            )

    def test_handles_multiline_injection(self, temp_dir):
        """Test injecting multiline HTML content."""
        template = temp_dir / 'template.html'
        template.write_text('<div><!-- ITEMS --></div>')

        multiline_content = '''<ul>
    <li>Item 1</li>
    <li>Item 2</li>
    <li>Item 3</li>
</ul>'''

        output = temp_dir / 'output.html'
        inject_content(template, output, {'ITEMS': multiline_content})

        result = output.read_text()
        assert '<li>Item 1</li>' in result
        assert '<li>Item 2</li>' in result
        assert '<li>Item 3</li>' in result


class TestValidateRequiredFields:
    """Test required field validation."""

    def test_returns_empty_for_valid_data(self):
        """Test that valid data returns no errors."""
        row = {'name': 'Alice', 'email': 'alice@example.com', 'age': 30}
        errors = validate_required_fields(row, ['name', 'email'], row_num=1)
        assert errors == []

    def test_detects_missing_field(self):
        """Test detection of completely missing field."""
        row = {'name': 'Alice'}
        errors = validate_required_fields(row, ['name', 'email'], row_num=1)
        assert len(errors) == 1
        assert 'email' in errors[0]
        assert 'Row 1' in errors[0]

    def test_detects_empty_string_field(self):
        """Test detection of empty string as missing."""
        row = {'name': 'Alice', 'email': ''}
        errors = validate_required_fields(row, ['name', 'email'], row_num=2)
        assert len(errors) == 1
        assert 'email' in errors[0]
        assert 'Row 2' in errors[0]

    def test_detects_whitespace_only_field(self):
        """Test detection of whitespace-only string as missing."""
        row = {'name': 'Alice', 'email': '   '}
        errors = validate_required_fields(row, ['name', 'email'], row_num=3)
        assert len(errors) == 1
        assert 'email' in errors[0]

    def test_detects_none_field(self):
        """Test detection of None value as missing."""
        row = {'name': 'Alice', 'email': None}
        errors = validate_required_fields(row, ['name', 'email'], row_num=1)
        assert len(errors) == 1
        assert 'email' in errors[0]

    def test_detects_multiple_missing_fields(self):
        """Test detection of multiple missing fields."""
        row = {'name': 'Alice'}
        errors = validate_required_fields(
            row, ['name', 'email', 'phone'], row_num=1
        )
        assert len(errors) == 2
        assert any('email' in e for e in errors)
        assert any('phone' in e for e in errors)

    def test_numeric_zero_is_valid(self):
        """Test that numeric zero is treated as valid value."""
        row = {'name': 'Test', 'count': 0}
        errors = validate_required_fields(row, ['name', 'count'], row_num=1)
        assert errors == []


class TestValidateUrlFormat:
    """Test URL format validation."""

    def test_valid_https_url(self):
        """Test that HTTPS URLs are valid."""
        assert validate_url_format('https://example.com') is True
        assert validate_url_format('https://example.com/path') is True
        assert validate_url_format('https://sub.example.com/path?q=1') is True

    def test_valid_http_url(self):
        """Test that HTTP URLs are valid."""
        assert validate_url_format('http://example.com') is True
        assert validate_url_format('http://localhost:8000') is True

    def test_invalid_urls(self):
        """Test that invalid URLs are rejected."""
        assert validate_url_format('example.com') is False
        assert validate_url_format('ftp://example.com') is False
        assert validate_url_format('not a url') is False
        assert validate_url_format('/relative/path') is False

    def test_empty_and_none(self):
        """Test that empty and None values are invalid."""
        assert validate_url_format('') is False
        assert validate_url_format(None) is False
        assert validate_url_format('   ') is False

    def test_handles_whitespace(self):
        """Test that URLs with surrounding whitespace are handled."""
        assert validate_url_format('  https://example.com  ') is True


class TestCheckFileExists:
    """Test file existence checking."""

    @pytest.fixture
    def temp_dir(self):
        """Create a real temporary directory with test files."""
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            # Create some test files
            (td_path / 'existing.png').touch()
            (td_path / 'another.jpg').touch()
            yield td_path

    def test_returns_none_for_existing_file(self, temp_dir):
        """Test that existing files return None (no error)."""
        result = check_file_exists('existing.png', temp_dir)
        assert result is None

    def test_returns_error_for_missing_file(self, temp_dir):
        """Test that missing files return an error message."""
        result = check_file_exists('nonexistent.png', temp_dir)
        assert result is not None
        assert 'not found' in result.lower()

    def test_empty_filename_returns_none(self, temp_dir):
        """Test that empty filename is OK (optional field)."""
        assert check_file_exists('', temp_dir) is None
        assert check_file_exists(None, temp_dir) is None
        assert check_file_exists('   ', temp_dir) is None

    def test_handles_path_with_subdirectory(self, temp_dir):
        """Test checking files in subdirectories."""
        subdir = temp_dir / 'subdir'
        subdir.mkdir()
        (subdir / 'file.txt').touch()

        result = check_file_exists('file.txt', subdir)
        assert result is None

        result = check_file_exists('missing.txt', subdir)
        assert result is not None


class TestIntegration:
    """Integration tests combining multiple utilities."""

    @pytest.fixture
    def temp_dir(self):
        """Create a real temporary directory."""
        with tempfile.TemporaryDirectory() as td:
            yield Path(td)

    def test_full_workflow_load_validate_inject(self, temp_dir):
        """Test complete workflow: load data, validate, inject into template."""
        # Create spreadsheet
        xlsx_path = temp_dir / 'data.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['title', 'author', 'year'])
        ws.append(['Paper One', 'Alice', 2024])
        ws.append(['Paper Two', 'Bob', 2023])
        wb.save(xlsx_path)

        # Create template
        template_path = temp_dir / 'template.html'
        template_path.write_text('''<!DOCTYPE html>
<html>
<body>
<h1>Papers</h1>
<!-- PAPERS -->
</body>
</html>''')

        # Load and validate
        rows = load_spreadsheet(xlsx_path)
        all_errors = []
        for i, row in enumerate(rows, start=2):
            errors = validate_required_fields(
                row, ['title', 'author', 'year'], i
            )
            all_errors.extend(errors)

        assert len(all_errors) == 0

        # Generate HTML content
        papers_html = '\n'.join([
            f'<div class="paper"><h2>{r["title"]}</h2>'
            f'<p>{r["author"]} ({r["year"]})</p></div>'
            for r in rows
        ])

        # Inject into template
        output_path = temp_dir / 'output.html'
        inject_content(template_path, output_path, {'PAPERS': papers_html})

        # Verify output
        result = output_path.read_text()
        assert 'Paper One' in result
        assert 'Paper Two' in result
        assert 'Alice' in result
        assert '2024' in result
