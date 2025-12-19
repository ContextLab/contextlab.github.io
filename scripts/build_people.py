#!/usr/bin/env python3
"""Build people.html from spreadsheet data.

Reads data from data/people.xlsx and generates people.html
using the template in templates/people.html.
"""
import re
from pathlib import Path
from typing import List, Dict, Any
import openpyxl

from utils import inject_content
from citation_utils import resolve_link


def parse_links_field(links_str: str) -> str:
    """Parse links field into HTML.

    Format: 'Label:URL, "Quoted Label":URL, ...'
    - Comma-separated pairs of label:url
    - Labels can be quoted for labels with spaces
    - URLs are resolved (local paths converted to GitHub URLs)

    Args:
        links_str: Links string in the format 'Label:URL, "Label":URL'

    Returns:
        HTML string like '[<a href="...">Label</a>] [<a href="...">Label</a>]'
    """
    if not links_str:
        return ''

    links = []
    # Parse the links string - handle both quoted and unquoted labels
    # Pattern: either "quoted label":url or label:url, separated by commas
    remaining = links_str.strip()

    while remaining:
        remaining = remaining.lstrip(' ,')
        if not remaining:
            break

        if remaining.startswith('"'):
            # Quoted label
            end_quote = remaining.find('"', 1)
            if end_quote == -1:
                break
            label = remaining[1:end_quote]
            rest = remaining[end_quote + 1:].lstrip()
            if rest.startswith(':'):
                rest = rest[1:]
            # Find end of URL (next comma or end of string)
            comma_pos = rest.find(',')
            if comma_pos == -1:
                url = rest.strip()
                remaining = ''
            else:
                url = rest[:comma_pos].strip()
                remaining = rest[comma_pos + 1:]
            links.append((label, url))
        else:
            # Unquoted label - split on first colon
            colon_pos = remaining.find(':')
            if colon_pos == -1:
                break
            label = remaining[:colon_pos].strip()
            rest = remaining[colon_pos + 1:]
            # Find end of URL - but URL may contain colons (https://)
            # So find the next comma that's not part of a URL
            comma_pos = rest.find(',')
            if comma_pos == -1:
                url = rest.strip()
                remaining = ''
            else:
                url = rest[:comma_pos].strip()
                remaining = rest[comma_pos + 1:]
            links.append((label, url))

    # Build HTML
    parts = []
    for label, url in links:
        if url:
            resolved_url = resolve_link(url, base_path="documents")
            parts.append(f'[<a href="{resolved_url}" target="_blank">{label}</a>]')

    return ' '.join(parts)


def load_people(xlsx_path: Path) -> Dict[str, List[Dict[str, Any]]]:
    """Load all people data from Excel spreadsheet.

    Args:
        xlsx_path: Path to the people.xlsx file

    Returns:
        Dictionary with keys for each section (director, members, etc.)
        Each value is a list of person dictionaries.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    data = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not any(cell is not None for cell in row):
                continue

            row_dict = {}
            for header, value in zip(headers, row):
                if value is None:
                    row_dict[header] = ''
                else:
                    row_dict[header] = value
            rows.append(row_dict)

        data[sheet_name] = rows

    wb.close()
    return data


def generate_director_content(director: Dict[str, Any]) -> str:
    """Generate HTML for the lab director section.

    Args:
        director: Dictionary with director data

    Returns:
        HTML string for director section
    """
    image = director.get('image', '')
    name = director.get('name', '')
    name_url = director.get('name_url', '')
    role = director.get('role', '')
    bio = director.get('bio', '')
    links_field = director.get('links_html', '')

    # Build image path (use placeholder if not specified)
    image_src = f"images/people/{image}" if image else "images/people/placeholder.png"

    # Build name with optional link
    if name_url:
        name_display = f'<a href="{name_url}" target="_blank">{name}</a>'
    else:
        name_display = name

    # Build role display
    role_display = f' | {role}' if role else ''

    # Parse links field into HTML
    links_html = parse_links_field(links_field)
    links_p = f'\n                    <p>{links_html}</p>' if links_html else ''

    html = f'''            <div class="two-column lab-director">
                <figure>
                    <img src="{image_src}" alt="{name}">
                </figure>
                <div>
                    <h3>{name_display}{role_display}</h3>
                    <p>{bio}</p>{links_p}
                </div>
            </div>'''

    return html


def generate_member_card(member: Dict[str, Any]) -> str:
    """Generate HTML for a single member card.

    Args:
        member: Dictionary with member data

    Returns:
        HTML string for the member card
    """
    image = member.get('image', '')
    name = member.get('name', '')
    name_url = member.get('name_url', '')
    role = member.get('role', '')
    bio = member.get('bio', '')

    # Build image path (use placeholder if not specified)
    image_src = f"images/people/{image}" if image else "images/people/placeholder.png"

    # Build name with optional link
    if name_url:
        name_display = f'<a href="{name_url}" target="_blank">{name}</a>'
    else:
        name_display = name

    # Build role display
    role_display = f' | {role}' if role else ''

    html = f'''                <div class="person-card">
                    <img src="{image_src}" alt="{name}">
                    <h3>{name_display}{role_display}</h3>
                    <p>{bio}</p>
                </div>'''

    return html


def generate_members_content(members: List[Dict[str, Any]]) -> str:
    """Generate HTML content for all active lab members.

    Members are arranged in a grid of 3 per row.

    Args:
        members: List of member dictionaries

    Returns:
        HTML string with all member cards organized in grids
    """
    if not members:
        return ''

    cards = [generate_member_card(m) for m in members]

    # Group cards into rows of 3
    grids = []
    for i in range(0, len(cards), 3):
        row_cards = cards[i:i+3]
        grid_html = '            <div class="people-grid">\n'
        grid_html += '\n'.join(row_cards)
        grid_html += '\n            </div>'
        grids.append(grid_html)

    return '\n\n'.join(grids)


def generate_alumni_entry(alum: Dict[str, Any]) -> str:
    """Generate HTML for a single alumni entry.

    Args:
        alum: Dictionary with alumni data (name, name_url, years, current_position, current_position_url)

    Returns:
        HTML string for the alumni entry
    """
    name = alum.get('name', '')
    name_url = alum.get('name_url', '')
    years = alum.get('years', '')
    current_position = alum.get('current_position', '')
    current_position_url = alum.get('current_position_url', '')

    # Build name with optional link
    if name_url:
        name_display = f'<a href="{name_url}" target="_blank">{name}</a>'
    else:
        name_display = name

    # Build position display with optional link
    # Position format is typically "now at Company" or "then a CDL grad student!"
    # We need to link the company/position name
    if current_position and current_position_url:
        # Extract the position name after "now at " or "then a "
        match = re.match(r'(now|then)\s+(at?)\s+(.+)', current_position)
        if match:
            prefix = f'{match.group(1)} {match.group(2)} '
            position_name = match.group(3)
            position_display = f'{prefix}<a href="{current_position_url}" target="_blank">{position_name}</a>'
        else:
            position_display = f'<a href="{current_position_url}" target="_blank">{current_position}</a>'
    else:
        position_display = current_position

    # Build parenthetical info
    paren_parts = []
    if years:
        paren_parts.append(years)
    if position_display:
        paren_parts.append(position_display)

    paren_display = f' ({"; ".join(paren_parts)})' if paren_parts else ''

    return f'{name_display}{paren_display}'


def generate_alumni_list_content(alumni: List[Dict[str, Any]]) -> str:
    """Generate HTML content for an alumni list (postdocs, grads, managers).

    Args:
        alumni: List of alumni dictionaries

    Returns:
        HTML string with alumni entries separated by <br>
    """
    if not alumni:
        return ''

    entries = [generate_alumni_entry(a) for a in alumni]
    return '<br>\n                    '.join(entries)


def generate_undergrad_entry(alum: Dict[str, Any]) -> str:
    """Generate HTML for a single undergraduate alumni entry.

    Args:
        alum: Dictionary with alumni data (name, years)

    Returns:
        HTML string for the alumni entry
    """
    name = alum.get('name', '')
    years = alum.get('years', '')

    if years:
        return f'{name} ({years})'
    return name


def generate_undergrad_list_content(alumni: List[Dict[str, Any]]) -> str:
    """Generate HTML content for undergraduate alumni list.

    Args:
        alumni: List of alumni dictionaries

    Returns:
        HTML string with alumni entries separated by <br>
    """
    if not alumni:
        return ''

    entries = [generate_undergrad_entry(a) for a in alumni]
    return '<br>\n                        '.join(entries)


def generate_collaborator_entry(collab: Dict[str, Any]) -> str:
    """Generate HTML for a single collaborator entry.

    Args:
        collab: Dictionary with collaborator data (name, url, description)

    Returns:
        HTML string for the collaborator paragraph
    """
    name = collab.get('name', '')
    url = collab.get('url', '')
    description = collab.get('description', '')

    # The description already contains the full text, but we need to replace
    # the name portion with a link
    if url:
        # If description starts with name, replace it with linked version
        if description.startswith(name):
            linked_name = f'<a href="{url}" target="_blank">{name}</a>'
            description = linked_name + description[len(name):]
        else:
            # Otherwise just create the link
            description = f'<a href="{url}" target="_blank">{name}</a>'

    return f'<p>{description}</p>'


def generate_collaborators_content(collaborators: List[Dict[str, Any]]) -> str:
    """Generate HTML content for collaborators section.

    Args:
        collaborators: List of collaborator dictionaries

    Returns:
        HTML string with all collaborator paragraphs
    """
    if not collaborators:
        return ''

    entries = [generate_collaborator_entry(c) for c in collaborators]
    return '\n                '.join(entries)


def build_people(
    data_path: Path,
    template_path: Path,
    output_path: Path
) -> None:
    """Build people.html from data and template.

    Args:
        data_path: Path to people.xlsx
        template_path: Path to template HTML file
        output_path: Path for generated HTML file
    """
    # Load data
    data = load_people(data_path)

    # Generate content for each section
    director_content = ''
    if data.get('director'):
        director_content = generate_director_content(data['director'][0])

    replacements = {
        'DIRECTOR_CONTENT': director_content,
        'MEMBERS_CONTENT': generate_members_content(data.get('members', [])),
        'ALUMNI_POSTDOCS_CONTENT': generate_alumni_list_content(data.get('alumni_postdocs', [])),
        'ALUMNI_GRADS_CONTENT': generate_alumni_list_content(data.get('alumni_grads', [])),
        'ALUMNI_MANAGERS_CONTENT': generate_alumni_list_content(data.get('alumni_managers', [])),
        'ALUMNI_UNDERGRADS_CONTENT': generate_undergrad_list_content(data.get('alumni_undergrads', [])),
        'COLLABORATORS_CONTENT': generate_collaborators_content(data.get('collaborators', [])),
    }

    # Inject into template
    inject_content(template_path, output_path, replacements)

    # Report
    total = sum(len(items) for items in data.values())
    print(f"Generated {output_path} with {total} people entries")


def main():
    """Main entry point for CLI usage."""
    project_root = Path(__file__).parent.parent
    data_path = project_root / 'data' / 'people.xlsx'
    template_path = project_root / 'templates' / 'people.html'
    output_path = project_root / 'people.html'

    build_people(data_path, template_path, output_path)


if __name__ == '__main__':
    main()
