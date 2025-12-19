#!/usr/bin/env python3
"""Sync mentored trainees between CV and people.xlsx.

Compares trainees in JRM_CV.tex with data/people.xlsx and generates
a sync report with recommendations. Can also apply updates.
"""
import re
from pathlib import Path
from typing import List, Dict, Any, Set, Tuple, Optional
from dataclasses import dataclass
import openpyxl
from openpyxl.utils import get_column_letter

from parse_cv_trainees import parse_cv_trainees, Trainee, get_active_trainees


@dataclass
class SyncAction:
    """Represents a sync action to take."""
    action: str  # 'add_to_spreadsheet', 'add_to_cv', 'update_spreadsheet', 'update_cv'
    target: str  # 'members', 'alumni_postdocs', 'alumni_grads', 'alumni_undergrads', 'cv'
    trainee: Trainee
    details: str


def normalize_name(name: str) -> str:
    """Normalize a name for comparison.

    Args:
        name: Name to normalize

    Returns:
        Lowercase name with extra whitespace removed
    """
    return ' '.join(name.lower().split())


# Common nickname mappings
NICKNAME_MAP = {
    'will': 'william',
    'bill': 'william',
    'billy': 'william',
    'charlie': 'charles',
    'chuck': 'charles',
    'alex': 'alexander',
    'alejandro': 'alexander',
    'maddy': 'madeline',
    'maddie': 'madeline',
    'chris': 'christopher',
    'mike': 'michael',
    'matt': 'matthew',
    'dan': 'daniel',
    'danny': 'daniel',
    'tom': 'thomas',
    'tommy': 'thomas',
    'bob': 'robert',
    'rob': 'robert',
    'dick': 'richard',
    'rick': 'richard',
    'jim': 'james',
    'jimmy': 'james',
    'joe': 'joseph',
    'joey': 'joseph',
    'sam': 'samuel',
    'sammy': 'samuel',
    'ben': 'benjamin',
    'benny': 'benjamin',
    'nick': 'nicholas',
    'nicky': 'nicholas',
    'tony': 'anthony',
    'dave': 'david',
    'davey': 'david',
    'ed': 'edward',
    'eddie': 'edward',
    'ted': 'theodore',
    'teddy': 'theodore',
    'steve': 'steven',
    'stevie': 'steven',
    'pete': 'peter',
    'petey': 'peter',
    'andy': 'andrew',
    'drew': 'andrew',
    'jake': 'jacob',
    'jay': 'jacob',
    'kate': 'katherine',
    'katy': 'katherine',
    'katie': 'katherine',
    'liz': 'elizabeth',
    'beth': 'elizabeth',
    'lizzy': 'elizabeth',
    'jenny': 'jennifer',
    'jen': 'jennifer',
    'jess': 'jessica',
    'jessie': 'jessica',
    'meg': 'margaret',
    'maggie': 'margaret',
    'peggy': 'margaret',
}


def expand_nicknames(name: str) -> Set[str]:
    """Expand a name to include nickname variations.

    Args:
        name: Name to expand

    Returns:
        Set of possible name variations
    """
    parts = name.lower().split()
    if not parts:
        return {name.lower()}

    first = parts[0]
    rest = ' '.join(parts[1:])

    variations = {name.lower()}

    # Add canonical form if first name is a nickname
    if first in NICKNAME_MAP:
        canonical = NICKNAME_MAP[first]
        if rest:
            variations.add(f"{canonical} {rest}")
        else:
            variations.add(canonical)

    # Add nickname forms if first name is a canonical form
    for nick, canon in NICKNAME_MAP.items():
        if first == canon:
            if rest:
                variations.add(f"{nick} {rest}")
            else:
                variations.add(nick)

    return variations


def names_match(name1: str, name2: str) -> bool:
    """Check if two names match (considering nicknames).

    Args:
        name1: First name
        name2: Second name

    Returns:
        True if names match
    """
    # Normalize both
    n1 = normalize_name(name1)
    n2 = normalize_name(name2)

    # Direct match
    if n1 == n2:
        return True

    # Check nickname expansions
    exp1 = expand_nicknames(n1)
    exp2 = expand_nicknames(n2)

    return bool(exp1 & exp2)  # Any overlap means match


def load_spreadsheet_names(xlsx_path: Path) -> Dict[str, Set[str]]:
    """Load all names from people.xlsx by sheet.

    Args:
        xlsx_path: Path to people.xlsx

    Returns:
        Dictionary mapping sheet name to set of normalized names
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    names = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_names = set()

        # Get header row to find name column
        headers = [cell.value for cell in sheet[1]]
        name_col = None
        for i, h in enumerate(headers):
            if h and 'name' in h.lower():
                name_col = i
                break

        if name_col is not None:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[name_col]:
                    sheet_names.add(normalize_name(str(row[name_col])))

        names[sheet_name] = sheet_names

    wb.close()
    return names


def load_spreadsheet_data(xlsx_path: Path) -> Dict[str, List[Dict[str, Any]]]:
    """Load all data from people.xlsx.

    Args:
        xlsx_path: Path to people.xlsx

    Returns:
        Dictionary mapping sheet name to list of row dictionaries
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    data = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        headers = [cell.value for cell in sheet[1]]

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(cell is not None for cell in row):
                continue
            row_dict = {}
            for header, value in zip(headers, row):
                row_dict[header] = value if value is not None else ''
            rows.append(row_dict)

        data[sheet_name] = rows

    wb.close()
    return data


def get_target_sheet(trainee: Trainee) -> str:
    """Determine which spreadsheet sheet a trainee belongs to.

    Args:
        trainee: Trainee object

    Returns:
        Sheet name ('members', 'alumni_postdocs', 'alumni_grads', 'alumni_undergrads')
    """
    if trainee.is_active:
        return 'members'

    # Alumni - route to appropriate sheet
    if trainee.category == 'postdoc':
        return 'alumni_postdocs'
    elif trainee.category == 'grad':
        return 'alumni_grads'
    else:  # undergrad
        return 'alumni_undergrads'


def format_years_for_spreadsheet(trainee: Trainee) -> str:
    """Format years for spreadsheet entry.

    Args:
        trainee: Trainee object

    Returns:
        Years string (e.g., '2021-2024' or '2021')
    """
    if trainee.start_year is None:
        return ''

    if trainee.end_year is None:
        # Still active - but this shouldn't be called for members sheet
        return str(trainee.start_year)

    if trainee.start_year == trainee.end_year:
        return str(trainee.start_year)

    return f"{trainee.start_year}-{trainee.end_year}"


def get_role_for_spreadsheet(trainee: Trainee) -> str:
    """Get role string for members spreadsheet.

    Args:
        trainee: Trainee object

    Returns:
        Role string
    """
    if trainee.category == 'postdoc':
        return 'postdoc'
    elif trainee.category == 'grad':
        if trainee.role:
            if 'doctoral' in trainee.role.lower():
                return 'grad student'
            elif 'masters' in trainee.role.lower():
                return 'masters student'
        return 'grad student'
    else:
        return 'undergrad'


def find_name_in_sheet(name: str, sheet_names: Set[str]) -> bool:
    """Check if a name (or nickname variant) is in a set of names.

    Args:
        name: Name to find
        sheet_names: Set of normalized names from a sheet

    Returns:
        True if name or a variant is found
    """
    for sheet_name in sheet_names:
        if names_match(name, sheet_name):
            return True
    return False


def find_name_match_in_names(name: str, all_names: Set[str]) -> Optional[str]:
    """Find a matching name in a set of names.

    Args:
        name: Name to find
        all_names: Set of names to search

    Returns:
        The matching name if found, None otherwise
    """
    for candidate in all_names:
        if names_match(name, candidate):
            return candidate
    return None


def compare_trainees(
    cv_trainees: Dict[str, List[Trainee]],
    spreadsheet_names: Dict[str, Set[str]],
    spreadsheet_data: Dict[str, List[Dict[str, Any]]]
) -> List[SyncAction]:
    """Compare CV trainees with spreadsheet and generate sync actions.

    Args:
        cv_trainees: Trainees parsed from CV
        spreadsheet_names: Names in spreadsheet by sheet
        spreadsheet_data: Full spreadsheet data

    Returns:
        List of SyncAction objects
    """
    actions = []

    # Collect all spreadsheet names (excluding director and collaborators)
    all_spreadsheet_names = set()
    for sheet_name, names in spreadsheet_names.items():
        if sheet_name not in ('director', 'collaborators'):
            all_spreadsheet_names.update(names)

    # Collect all CV names (with their canonical forms for matching)
    all_cv_names = set()
    cv_name_to_trainee = {}
    for category, trainees in cv_trainees.items():
        for trainee in trainees:
            norm_name = normalize_name(trainee.name)
            all_cv_names.add(norm_name)
            cv_name_to_trainee[norm_name] = trainee

    # Check each CV trainee
    for category, trainees in cv_trainees.items():
        for trainee in trainees:
            norm_name = normalize_name(trainee.name)
            target_sheet = get_target_sheet(trainee)

            # Check if in the correct sheet (using nickname matching)
            in_target = find_name_in_sheet(norm_name, spreadsheet_names.get(target_sheet, set()))

            if not in_target:
                # Not in correct sheet - check if in any sheet
                found_in = None
                for sheet_name, names in spreadsheet_names.items():
                    if sheet_name in ('director', 'collaborators'):
                        continue
                    if find_name_in_sheet(norm_name, names):
                        found_in = sheet_name
                        break

                if found_in:
                    # In wrong sheet - but this might be expected
                    # (e.g., Paxton is both a grad student AND former undergrad)
                    # Only flag if it's a clear error (active person in alumni sheet)
                    if trainee.is_active and found_in.startswith('alumni'):
                        actions.append(SyncAction(
                            action='move_in_spreadsheet',
                            target=target_sheet,
                            trainee=trainee,
                            details=f"Move from {found_in} to {target_sheet}"
                        ))
                else:
                    # Not in spreadsheet at all
                    actions.append(SyncAction(
                        action='add_to_spreadsheet',
                        target=target_sheet,
                        trainee=trainee,
                        details=f"Add to {target_sheet}"
                    ))

    # Check for spreadsheet entries not in CV
    # Only check members, alumni_postdocs, alumni_grads, alumni_undergrads
    check_sheets = ['members', 'alumni_postdocs', 'alumni_grads', 'alumni_undergrads']
    for sheet_name in check_sheets:
        for name in spreadsheet_names.get(sheet_name, set()):
            # Check if this name matches any CV name (using nickname matching)
            match = find_name_match_in_names(name, all_cv_names)
            if match is None:
                # Get the row data
                row_data = None
                for row in spreadsheet_data.get(sheet_name, []):
                    row_name = row.get('name', '')
                    if normalize_name(str(row_name)) == name:
                        row_data = row
                        break

                # Create a placeholder trainee for reporting
                placeholder = Trainee(
                    name=name.title(),  # Title case for display
                    category='unknown',
                    role=row_data.get('role', '') if row_data else '',
                    start_year=None,
                    end_year=None,
                    current_position=None
                )

                actions.append(SyncAction(
                    action='add_to_cv',
                    target='cv',
                    trainee=placeholder,
                    details=f"Found in {sheet_name} but not in CV"
                ))

    return actions


def print_sync_report(actions: List[SyncAction]) -> None:
    """Print a formatted sync report.

    Args:
        actions: List of sync actions
    """
    print("=" * 60)
    print("CV <-> SPREADSHEET SYNC REPORT")
    print("=" * 60)

    if not actions:
        print("\nNo sync actions needed - CV and spreadsheet are in sync!")
        return

    # Group by action type
    add_to_ss = [a for a in actions if a.action == 'add_to_spreadsheet']
    add_to_cv = [a for a in actions if a.action == 'add_to_cv']
    move_in_ss = [a for a in actions if a.action == 'move_in_spreadsheet']

    if add_to_ss:
        print(f"\n--- ADD TO SPREADSHEET ({len(add_to_ss)} entries) ---")
        # Group by target sheet
        by_sheet = {}
        for a in add_to_ss:
            by_sheet.setdefault(a.target, []).append(a)

        for sheet, sheet_actions in sorted(by_sheet.items()):
            print(f"\n  [{sheet}]")
            for a in sheet_actions:
                years = format_years_for_spreadsheet(a.trainee)
                role = get_role_for_spreadsheet(a.trainee)
                pos = f" -> {a.trainee.current_position}" if a.trainee.current_position else ""
                print(f"    - {a.trainee.name} ({role}, {years}){pos}")

    if add_to_cv:
        print(f"\n--- ADD TO CV ({len(add_to_cv)} entries) ---")
        for a in add_to_cv:
            print(f"    - {a.trainee.name}: {a.details}")

    if move_in_ss:
        print(f"\n--- MOVE WITHIN SPREADSHEET ({len(move_in_ss)} entries) ---")
        for a in move_in_ss:
            print(f"    - {a.trainee.name}: {a.details}")

    print("\n" + "=" * 60)


def apply_spreadsheet_updates(
    xlsx_path: Path,
    actions: List[SyncAction],
    dry_run: bool = True
) -> None:
    """Apply sync actions to the spreadsheet.

    Args:
        xlsx_path: Path to people.xlsx
        actions: List of sync actions
        dry_run: If True, don't actually save changes
    """
    add_actions = [a for a in actions if a.action == 'add_to_spreadsheet']

    if not add_actions:
        print("No spreadsheet additions needed.")
        return

    wb = openpyxl.load_workbook(xlsx_path)

    for action in add_actions:
        sheet_name = action.target
        trainee = action.trainee

        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet {sheet_name} not found")
            continue

        ws = wb[sheet_name]

        # Get headers
        headers = [cell.value for cell in ws[1]]

        # Build new row based on sheet type
        if sheet_name == 'members':
            # Members: image, name, name_url, role, bio, links_html
            new_row = [
                '',  # image - needs to be added manually
                trainee.name.lower(),
                '',  # name_url
                get_role_for_spreadsheet(trainee),
                '',  # bio - needs to be added manually
                ''   # links_html
            ]
        elif sheet_name in ('alumni_postdocs', 'alumni_grads'):
            # Alumni: name, name_url, years, current_position, current_position_url
            years = format_years_for_spreadsheet(trainee)
            if sheet_name == 'alumni_grads' and trainee.role:
                years = f"{trainee.role}, {years}"
            new_row = [
                trainee.name,
                '',  # name_url
                years,
                f"now at {trainee.current_position}" if trainee.current_position else '',
                ''   # current_position_url
            ]
        elif sheet_name == 'alumni_undergrads':
            # Undergrad alumni: name, years
            new_row = [
                trainee.name,
                format_years_for_spreadsheet(trainee)
            ]
        else:
            print(f"Warning: Unknown sheet type {sheet_name}")
            continue

        # Append row
        ws.append(new_row)
        print(f"{'[DRY RUN] ' if dry_run else ''}Added {trainee.name} to {sheet_name}")

    if not dry_run:
        wb.save(xlsx_path)
        print(f"\nSaved changes to {xlsx_path}")
    else:
        print(f"\n[DRY RUN] Would save changes to {xlsx_path}")

    wb.close()


def sort_members_sheet(xlsx_path: Path, dry_run: bool = True) -> None:
    """Sort the members sheet by category and start year.

    Order: postdoc > grad student > lab manager > research scientist > undergrad
    Within category: reverse chronological by start year

    Args:
        xlsx_path: Path to people.xlsx
        dry_run: If True, don't actually save changes
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['members']

    # Get headers
    headers = [cell.value for cell in ws[1]]

    # Read all data rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            rows.append(list(row))

    # Define category order
    category_order = {
        'postdoc': 0,
        'grad student': 1,
        'masters student': 2,
        'lab manager': 3,
        'research scientist': 4,
        'undergrad': 5,
    }

    def get_sort_key(row):
        role_idx = headers.index('role') if 'role' in headers else 3
        role = (row[role_idx] or '').lower()
        cat_order = category_order.get(role, 99)

        # For year, we'd need to parse from bio or add a start_year column
        # For now, just sort by category
        return (cat_order, row[1] or '')  # Then by name

    rows.sort(key=get_sort_key)

    # Clear existing data and rewrite
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).value = None

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    if not dry_run:
        wb.save(xlsx_path)
        print(f"Sorted members sheet in {xlsx_path}")
    else:
        print(f"[DRY RUN] Would sort members sheet in {xlsx_path}")

    wb.close()


def generate_cv_additions(actions: List[SyncAction]) -> str:
    """Generate LaTeX entries for people to add to CV.

    Args:
        actions: List of sync actions

    Returns:
        LaTeX text to add to CV
    """
    add_to_cv = [a for a in actions if a.action == 'add_to_cv']

    if not add_to_cv:
        return ""

    lines = ["% === People to add to CV ==="]
    for action in add_to_cv:
        lines.append(f"% {action.trainee.name}: {action.details}")

    return "\n".join(lines)


def main():
    """Main entry point."""
    import argparse

    parser = argparse.ArgumentParser(description='Sync CV trainees with people.xlsx')
    parser.add_argument('--apply', action='store_true', help='Apply changes (default: dry run)')
    parser.add_argument('--sort', action='store_true', help='Sort members sheet')
    args = parser.parse_args()

    project_root = Path(__file__).parent.parent
    cv_path = project_root / 'documents' / 'JRM_CV.tex'
    xlsx_path = project_root / 'data' / 'people.xlsx'

    # Parse CV
    print("Parsing CV...")
    cv_trainees = parse_cv_trainees(cv_path)

    # Load spreadsheet
    print("Loading spreadsheet...")
    spreadsheet_names = load_spreadsheet_names(xlsx_path)
    spreadsheet_data = load_spreadsheet_data(xlsx_path)

    # Compare
    print("Comparing...")
    actions = compare_trainees(cv_trainees, spreadsheet_names, spreadsheet_data)

    # Print report
    print_sync_report(actions)

    # Apply if requested
    if args.apply:
        apply_spreadsheet_updates(xlsx_path, actions, dry_run=False)
        if args.sort:
            sort_members_sheet(xlsx_path, dry_run=False)
    else:
        print("\nRun with --apply to make changes")

    # Print CV additions if any
    cv_additions = generate_cv_additions(actions)
    if cv_additions:
        print("\n" + cv_additions)


if __name__ == '__main__':
    main()
