#!/usr/bin/env python3
"""Extract people data from existing HTML into Excel spreadsheet.

This is a one-time script to migrate existing HTML data to the spreadsheet format.
"""
import re
from pathlib import Path
from bs4 import BeautifulSoup
import openpyxl


def extract_people(html_path: Path) -> dict:
    """Extract all people data from HTML file.

    Returns dict with keys: director, members, alumni_postdocs, alumni_grads,
    alumni_managers, alumni_undergrads, collaborators
    """
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f.read(), 'html.parser')

    data = {}

    # Extract lab director
    director_section = soup.find('div', class_='lab-director')
    if director_section:
        img = director_section.find('img')
        h3 = director_section.find('h3')
        ps = director_section.find_all('p')

        name_parts = h3.get_text().split('|') if h3 else ['', '']
        name = name_parts[0].strip() if name_parts else ''
        role = name_parts[1].strip() if len(name_parts) > 1 else ''

        data['director'] = [{
            'image': img.get('src', '').replace('images/people/', '') if img else '',
            'name': name,
            'role': role,
            'bio': ps[0].get_text(strip=True) if ps else '',
            'links_html': get_inner_html(ps[1]) if len(ps) > 1 else '',
            'name_url': ''
        }]
        print(f"Extracted director: {name}")

    # Extract active lab members
    members = []
    people_grids = soup.find_all('div', class_='people-grid')
    for grid in people_grids:
        cards = grid.find_all('div', class_='person-card')
        for card in cards:
            member = extract_person_card(card)
            members.append(member)

    data['members'] = members
    print(f"Extracted {len(members)} active lab members")

    # Extract alumni (from the lab-alumni section)
    alumni_section = soup.find('section', id='lab-alumni')
    if alumni_section:
        # Find all h3 headers that identify categories
        h3s = alumni_section.find_all('h3')
        for h3 in h3s:
            header_text = h3.get_text(strip=True).lower()
            next_elem = h3.find_next_sibling()

            # Skip non-alumni headers
            if 'who we were' in header_text or 'fledgling' in header_text.lower():
                continue

            if 'postdoctoral' in header_text:
                data['alumni_postdocs'] = extract_alumni_list(next_elem)
                print(f"Extracted {len(data['alumni_postdocs'])} postdocs")
            elif 'undergraduate' in header_text:
                # Check undergraduate BEFORE graduate since 'undergraduate' contains 'graduate'
                alumni_list = next_elem
                if alumni_list and alumni_list.name == 'p':
                    data['alumni_undergrads'] = extract_alumni_simple_list(alumni_list)
                else:
                    data['alumni_undergrads'] = []
                print(f"Extracted {len(data.get('alumni_undergrads', []))} undergrad alumni")
            elif 'graduate' in header_text:
                data['alumni_grads'] = extract_alumni_list(next_elem)
                print(f"Extracted {len(data['alumni_grads'])} grad students")
            elif 'manager' in header_text:
                data['alumni_managers'] = extract_alumni_list(next_elem)
                print(f"Extracted {len(data['alumni_managers'])} lab managers")

    # Extract collaborators
    collab_section = soup.find('section', id='collaborators')
    if collab_section:
        collaborators = []
        # Find all paragraphs with links
        for p in collab_section.find_all('p'):
            link = p.find('a')
            if link and p.get_text(strip=True):
                collab = {
                    'name': link.get_text(strip=True),
                    'url': link.get('href', ''),
                    'description': p.get_text(strip=True)
                }
                collaborators.append(collab)

        data['collaborators'] = collaborators
        print(f"Extracted {len(collaborators)} collaborators")

    return data


def extract_person_card(card) -> dict:
    """Extract data from a person card."""
    img = card.find('img')
    h3 = card.find('h3')
    p = card.find('p')

    # Parse name and role from h3
    name = ''
    role = ''
    name_url = ''

    if h3:
        link = h3.find('a')
        if link:
            name_url = link.get('href', '')

        text = h3.get_text()
        if '|' in text:
            parts = text.split('|')
            name = parts[0].strip()
            role = parts[1].strip()
        else:
            name = text.strip()

    return {
        'image': img.get('src', '').replace('images/people/', '') if img else '',
        'name': name,
        'name_url': name_url,
        'role': role,
        'bio': p.get_text(strip=True) if p else '',
        'links_html': ''
    }


def extract_alumni_list(elem) -> list:
    """Extract alumni from a paragraph with <br> separated entries.

    Handles formats like:
    - Gina Notaro (2017-2018; now at <a>Lockheed Martin</a>)
    - <a>Andrew Heusser</a> (2016-2018; now at <a>Akili</a>)
    """
    alumni = []
    if not elem:
        return alumni

    # Get the raw HTML and split by <br>
    html = str(elem)
    entries = re.split(r'<br\s*/?>', html)

    for entry in entries:
        entry = entry.strip()
        if not entry:
            continue
        # Remove opening/closing <p> tags but keep the content
        entry = re.sub(r'^<p[^>]*>', '', entry)
        entry = re.sub(r'</p>$', '', entry)
        entry = entry.strip()
        if not entry:
            continue

        entry_soup = BeautifulSoup(entry, 'html.parser')
        full_text = entry_soup.get_text(strip=True)
        if not full_text:
            continue

        alum = {
            'name': '',
            'name_url': '',
            'years': '',
            'current_position': '',
            'current_position_url': ''
        }

        # Find all links in this entry
        links = entry_soup.find_all('a')

        # The first link (if any) that appears BEFORE the parenthesis is the person's name
        # Links inside parenthesis are their current position
        paren_pos = full_text.find('(')

        # Determine which link is the name vs current position
        name_link = None
        position_link = None

        for link in links:
            link_text = link.get_text(strip=True)
            link_pos_in_text = full_text.find(link_text)
            if paren_pos < 0 or link_pos_in_text < paren_pos:
                # Link is before parenthesis - this is the name
                if name_link is None:
                    name_link = link
            else:
                # Link is inside parenthesis - this is the current position
                position_link = link

        # Get name - either from link or from text before parenthesis
        if name_link:
            alum['name'] = name_link.get_text(strip=True)
            alum['name_url'] = name_link.get('href', '')
        else:
            # No name link, name is text before (
            if paren_pos > 0:
                alum['name'] = full_text[:paren_pos].strip()
            else:
                alum['name'] = full_text.strip()

        # Get position link URL if available
        if position_link:
            alum['current_position_url'] = position_link.get('href', '')

        # Parse years and current position from parenthesis
        paren_match = re.search(r'\(([^)]+)\)', full_text)
        if paren_match:
            paren_content = paren_match.group(1)
            if ';' in paren_content:
                parts = paren_content.split(';', 1)
                alum['years'] = parts[0].strip()
                # Clean up spacing issues from stripped links
                position = parts[1].strip()
                position = re.sub(r'(now|then)\s+(at?)\s*', r'\1 \2 ', position)
                alum['current_position'] = position
            else:
                alum['years'] = paren_content.strip()

        if alum['name']:
            alumni.append(alum)

    return alumni


def extract_alumni_simple_list(elem) -> list:
    """Extract simple alumni list (just names and years)."""
    alumni = []
    if not elem:
        return alumni

    # Get text and split by <br>
    html = str(elem)
    entries = re.split(r'<br\s*/?>', html)

    for entry in entries:
        entry_soup = BeautifulSoup(entry, 'html.parser')
        text = entry_soup.get_text(strip=True)
        if not text:
            continue

        # Format: "Name (years)"
        match = re.match(r'(.+?)\s*\(([^)]+)\)', text)
        if match:
            alumni.append({
                'name': match.group(1).strip(),
                'years': match.group(2).strip()
            })
        else:
            alumni.append({
                'name': text.strip(),
                'years': ''
            })

    return alumni


def get_inner_html(element) -> str:
    """Get the inner HTML of an element as a string."""
    if not element:
        return ''
    return ''.join(str(child) for child in element.children).strip()


def save_to_excel(data: dict, output_path: Path):
    """Save extracted data to Excel spreadsheet with multiple sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Director sheet
    ws = wb.create_sheet(title='director')
    ws.append(['image', 'name', 'name_url', 'role', 'bio', 'links_html'])
    for item in data.get('director', []):
        ws.append([item.get('image', ''), item.get('name', ''), item.get('name_url', ''),
                   item.get('role', ''), item.get('bio', ''), item.get('links_html', '')])

    # Active members sheet
    ws = wb.create_sheet(title='members')
    ws.append(['image', 'name', 'name_url', 'role', 'bio', 'links_html'])
    for item in data.get('members', []):
        ws.append([item.get('image', ''), item.get('name', ''), item.get('name_url', ''),
                   item.get('role', ''), item.get('bio', ''), item.get('links_html', '')])

    # Alumni postdocs sheet
    ws = wb.create_sheet(title='alumni_postdocs')
    ws.append(['name', 'name_url', 'years', 'current_position', 'current_position_url'])
    for item in data.get('alumni_postdocs', []):
        ws.append([item.get('name', ''), item.get('name_url', ''),
                   item.get('years', ''), item.get('current_position', ''),
                   item.get('current_position_url', '')])

    # Alumni grads sheet
    ws = wb.create_sheet(title='alumni_grads')
    ws.append(['name', 'name_url', 'years', 'current_position', 'current_position_url'])
    for item in data.get('alumni_grads', []):
        ws.append([item.get('name', ''), item.get('name_url', ''),
                   item.get('years', ''), item.get('current_position', ''),
                   item.get('current_position_url', '')])

    # Alumni managers sheet
    ws = wb.create_sheet(title='alumni_managers')
    ws.append(['name', 'name_url', 'years', 'current_position', 'current_position_url'])
    for item in data.get('alumni_managers', []):
        ws.append([item.get('name', ''), item.get('name_url', ''),
                   item.get('years', ''), item.get('current_position', ''),
                   item.get('current_position_url', '')])

    # Alumni undergrads sheet
    ws = wb.create_sheet(title='alumni_undergrads')
    ws.append(['name', 'years'])
    for item in data.get('alumni_undergrads', []):
        ws.append([item.get('name', ''), item.get('years', '')])

    # Collaborators sheet
    ws = wb.create_sheet(title='collaborators')
    ws.append(['name', 'url', 'description'])
    for item in data.get('collaborators', []):
        ws.append([item.get('name', ''), item.get('url', ''), item.get('description', '')])

    # Adjust column widths
    for sheet in wb.worksheets:
        for col in range(1, 7):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 40

    wb.save(output_path)
    print(f"Saved to {output_path}")


def main():
    project_root = Path(__file__).parent.parent
    html_path = project_root / 'people.html'
    output_path = project_root / 'data' / 'people.xlsx'

    print(f"Extracting from: {html_path}")
    data = extract_people(html_path)

    save_to_excel(data, output_path)
    print("Done!")


if __name__ == '__main__':
    main()
