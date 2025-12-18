#!/usr/bin/env python3
"""Fix the members sheet to match the current lab roster.

This script updates the members sheet based on the actual current
lab members list and moves others to alumni.
"""
from pathlib import Path
import openpyxl

# Current active members (as of Dec 2024)
CURRENT_MEMBERS = {
    # Grad students
    'claudia gonciulea': {'role': 'grad student', 'category': 'grad'},
    'paxton fitzpatrick': {'role': 'grad student', 'category': 'grad'},
    'xinming xu': {'role': 'grad student', 'category': 'grad'},
    # Undergrads
    'aidan miller': {'role': 'undergrad', 'category': 'undergrad', 'new': True},
    'alexandra wingo': {'role': 'undergrad', 'category': 'undergrad'},
    'alishba tahir': {'role': 'undergrad', 'category': 'undergrad', 'new': True},
    'angelyn liu': {'role': 'undergrad', 'category': 'undergrad'},
    'annabelle morrow': {'role': 'undergrad', 'category': 'undergrad'},
    'azaire andre': {'role': 'undergrad', 'category': 'undergrad', 'new': True},
    'ellie mattox': {'role': 'undergrad', 'category': 'undergrad', 'new': True},
    'emmy thornton': {'role': 'undergrad', 'category': 'undergrad', 'new': True},
    'evan mcdermid': {'role': 'undergrad', 'category': 'undergrad', 'new': True},
    'jackson c. sandrich': {'role': 'undergrad', 'category': 'undergrad', 'new': True},
    'jacob bacus': {'role': 'undergrad', 'category': 'undergrad'},
    'jennifer xu': {'role': 'undergrad', 'category': 'undergrad'},
    'luca gandrud': {'role': 'undergrad', 'category': 'undergrad', 'new': True},
    'om shah': {'role': 'undergrad', 'category': 'undergrad'},
    'sam haskel': {'role': 'undergrad', 'category': 'undergrad'},
    'sarah parigela': {'role': 'undergrad', 'category': 'undergrad'},
    'will lehman': {'role': 'undergrad', 'category': 'undergrad', 'new': True},
    # Active but not in initial list
    'kevin chang': {'role': 'undergrad', 'category': 'undergrad'},
    'andrew richardson': {'role': 'undergrad', 'category': 'undergrad'},
    'ben hanson': {'role': 'undergrad', 'category': 'undergrad'},
    'owen phillips': {'role': 'undergrad', 'category': 'undergrad'},
    'joy maina': {'role': 'undergrad', 'category': 'undergrad'},
    'chelsea joe': {'role': 'undergrad', 'category': 'undergrad'},
    'jake mcdermid': {'role': 'undergrad', 'category': 'undergrad'},
}


def normalize_name(name):
    """Normalize a name for comparison."""
    return ' '.join(str(name).lower().split())


def main():
    project_root = Path(__file__).parent.parent
    xlsx_path = project_root / 'data' / 'people.xlsx'

    wb = openpyxl.load_workbook(xlsx_path)

    # Get current members sheet data
    ws_members = wb['members']
    ws_alumni = wb['alumni_undergrads']

    # Get headers
    member_headers = [cell.value for cell in ws_members[1]]
    alumni_headers = [cell.value for cell in ws_alumni[1]]

    # Read current members
    current_rows = []
    rows_to_move_to_alumni = []

    for row_idx in range(2, ws_members.max_row + 1):
        row_data = [ws_members.cell(row=row_idx, column=col).value for col in range(1, len(member_headers) + 1)]
        if not any(row_data):
            continue

        name = normalize_name(row_data[1]) if row_data[1] else ''

        if name in CURRENT_MEMBERS:
            # Keep this member, update role if needed
            row_data[3] = CURRENT_MEMBERS[name]['role']
            current_rows.append(row_data)
            print(f"Keeping: {name}")
        elif name:
            # Move to alumni
            rows_to_move_to_alumni.append({
                'name': row_data[1],
                'years': '2024'  # Approximate end year
            })
            print(f"Moving to alumni: {name}")

    # Add new members that aren't in the spreadsheet yet
    existing_names = {normalize_name(r[1]) for r in current_rows if r[1]}
    for name, info in CURRENT_MEMBERS.items():
        if name not in existing_names:
            new_row = [
                '',  # image
                name,  # name (will be title-cased later)
                '',  # name_url
                info['role'],
                '',  # bio
                ''   # links_html
            ]
            current_rows.append(new_row)
            print(f"Adding new member: {name}")

    # Clear members sheet (except header)
    for row_idx in range(2, ws_members.max_row + 1):
        for col_idx in range(1, len(member_headers) + 1):
            ws_members.cell(row=row_idx, column=col_idx).value = None

    # Sort members: grad students first, then undergrads
    def sort_key(row):
        role = (row[3] or '').lower()
        name = (row[1] or '').lower()
        if 'grad' in role:
            return (0, name)
        return (1, name)

    current_rows.sort(key=sort_key)

    # Write back members
    for row_idx, row_data in enumerate(current_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_members.cell(row=row_idx, column=col_idx).value = value

    # Add moved members to alumni_undergrads
    next_alumni_row = ws_alumni.max_row + 1
    for alum in rows_to_move_to_alumni:
        ws_alumni.cell(row=next_alumni_row, column=1).value = alum['name']
        ws_alumni.cell(row=next_alumni_row, column=2).value = alum['years']
        next_alumni_row += 1
        print(f"Added to alumni_undergrads: {alum['name']}")

    # Save
    wb.save(xlsx_path)
    print(f"\nSaved changes to {xlsx_path}")

    # Print summary
    print(f"\n=== Summary ===")
    print(f"Active members: {len(current_rows)}")
    print(f"Moved to alumni: {len(rows_to_move_to_alumni)}")

    # Print new members that need CV additions
    new_members = [name for name, info in CURRENT_MEMBERS.items() if info.get('new')]
    if new_members:
        print(f"\n=== New members to add to CV ===")
        for name in sorted(new_members):
            print(f"  - {name.title()}")

    wb.close()


if __name__ == '__main__':
    main()
