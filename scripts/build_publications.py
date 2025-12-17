#!/usr/bin/env python3
"""Build publications.html from spreadsheet data.

Reads data from data/publications.xlsx and generates publications.html
using the template in templates/publications.html.

The spreadsheet uses structured fields for each publication type:
- papers: authors, year, title, journal, volume, issue, pages, status, links
- chapters: authors, year, title, editors, book_title, publisher info, links
- dissertations: authors, year, title, degree_type, institution, location, links
- talks: authors, year, title, venue info, links
- courses: title, description, links
- posters: authors, year, title, conference, location, session_number, links
"""
from pathlib import Path
from typing import List, Dict, Any
import openpyxl

from utils import inject_content
from citation_utils import (
    format_paper_citation,
    format_chapter_citation,
    format_dissertation_citation,
    format_talk_citation,
    format_poster_citation,
    format_course_citation,
    markdown_to_html,
    resolve_link,
    build_links_html,
)


def load_publications(xlsx_path: Path) -> Dict[str, List[Dict[str, Any]]]:
    """Load all publication data from Excel spreadsheet.

    Args:
        xlsx_path: Path to the publications.xlsx file

    Returns:
        Dictionary with keys for each section (papers, chapters, etc.)
        Each value is a list of publication dictionaries.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    data = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not any(cell is not None for cell in row):
                continue

            row_dict = {}
            for header, value in zip(headers, row):
                if value is None:
                    row_dict[header] = ''
                else:
                    row_dict[header] = str(value) if not isinstance(value, str) else value
            rows.append(row_dict)

        data[sheet_name] = rows

    wb.close()
    return data


def build_paper_citation(pub: Dict[str, Any]) -> str:
    """Build citation HTML for a paper."""
    return format_paper_citation(
        authors=pub.get('authors', ''),
        year=pub.get('year', ''),
        title=pub.get('title', ''),
        journal=pub.get('journal', ''),
        volume=pub.get('volume', ''),
        issue=pub.get('issue', ''),
        pages=pub.get('pages', ''),
        status=pub.get('status', ''),
        preprint_id=pub.get('preprint_id', '')
    )


def build_chapter_citation(pub: Dict[str, Any]) -> str:
    """Build citation HTML for a book chapter."""
    return format_chapter_citation(
        authors=pub.get('authors', ''),
        year=pub.get('year', ''),
        title=pub.get('title', ''),
        editors=pub.get('editors', ''),
        book_title=pub.get('book_title', ''),
        publisher_location=pub.get('publisher_location', ''),
        publisher=pub.get('publisher', '')
    )


def build_dissertation_citation(pub: Dict[str, Any]) -> str:
    """Build citation HTML for a dissertation."""
    return format_dissertation_citation(
        authors=pub.get('authors', ''),
        year=pub.get('year', ''),
        title=pub.get('title', ''),
        degree_type=pub.get('degree_type', ''),
        institution=pub.get('institution', ''),
        location=pub.get('location', '')
    )


def build_talk_citation(pub: Dict[str, Any]) -> str:
    """Build citation HTML for a talk."""
    return format_talk_citation(
        authors=pub.get('authors', ''),
        year=pub.get('year', ''),
        title=pub.get('title', ''),
        venue_name=pub.get('venue_name', ''),
        venue_url=pub.get('venue_url', '')
    )


def build_poster_citation(pub: Dict[str, Any]) -> str:
    """Build citation HTML for a poster."""
    return format_poster_citation(
        authors=pub.get('authors', ''),
        year=pub.get('year', ''),
        title=pub.get('title', ''),
        conference=pub.get('conference', ''),
        location=pub.get('location', ''),
        session_number=pub.get('session_number', '')
    )


def build_course_citation(pub: Dict[str, Any]) -> str:
    """Build citation HTML for a course (just the description)."""
    description = pub.get('description', '')
    # Convert markdown to HTML in description
    return markdown_to_html(description)


def parse_extra_links(extra_links: str) -> List[tuple]:
    """Parse extra_links field into list of (label, url) tuples.

    Format: "Label1:URL1;Label2:URL2"
    """
    if not extra_links:
        return []

    links = []
    for part in extra_links.split(';'):
        if ':' in part:
            # Split on first colon only (URLs contain colons)
            label, url = part.split(':', 1)
            links.append((label.strip(), url.strip()))
    return links


def build_links_for_paper(pub: Dict[str, Any]) -> str:
    """Build links HTML for a paper."""
    links = []
    if pub.get('pdf_link'):
        links.append(('PDF', pub['pdf_link']))
    if pub.get('code_link'):
        # Check if it's combined CODE + DATA
        if pub.get('data_link') == pub.get('code_link'):
            links.append(('CODE + DATA', pub['code_link']))
        else:
            links.append(('CODE', pub['code_link']))
            if pub.get('data_link'):
                links.append(('DATA', pub['data_link']))
    elif pub.get('data_link'):
        links.append(('DATA', pub['data_link']))

    # Add extra links
    links.extend(parse_extra_links(pub.get('extra_links', '')))

    return build_links_html(links)


def build_links_for_chapter(pub: Dict[str, Any]) -> str:
    """Build links HTML for a chapter."""
    links = []
    if pub.get('pdf_link'):
        links.append(('PDF', pub['pdf_link']))
    links.extend(parse_extra_links(pub.get('extra_links', '')))
    return build_links_html(links)


def build_links_for_dissertation(pub: Dict[str, Any]) -> str:
    """Build links HTML for a dissertation."""
    links = []
    if pub.get('pdf_link'):
        links.append(('PDF', pub['pdf_link']))
    links.extend(parse_extra_links(pub.get('extra_links', '')))
    return build_links_html(links)


def build_links_for_talk(pub: Dict[str, Any]) -> str:
    """Build links HTML for a talk."""
    links = []
    if pub.get('paper_link'):
        links.append(('PAPER', pub['paper_link']))
    if pub.get('code_link'):
        links.append(('CODE', pub['code_link']))
    if pub.get('data_link'):
        links.append(('DATA', pub['data_link']))
    links.extend(parse_extra_links(pub.get('extra_links', '')))
    return build_links_html(links)


def build_links_for_course(pub: Dict[str, Any]) -> str:
    """Build links HTML for a course."""
    links = []
    if pub.get('github_link'):
        links.append(('GitHub', pub['github_link']))
    links.extend(parse_extra_links(pub.get('extra_links', '')))
    return build_links_html(links)


def build_links_for_poster(pub: Dict[str, Any]) -> str:
    """Build links HTML for a poster.

    Note: Posters don't show separate PDF links - the title itself links to the PDF.
    Only extra_links are shown if present.
    """
    links = []
    links.extend(parse_extra_links(pub.get('extra_links', '')))
    return build_links_html(links)


def generate_publication_card(pub: Dict[str, Any], pub_type: str) -> str:
    """Generate HTML for a single publication card.

    Args:
        pub: Dictionary with publication data
        pub_type: Type of publication (papers, chapters, etc.)

    Returns:
        HTML string for the publication card
    """
    image = pub.get('image', '')
    title = pub.get('title', '')
    title_url = pub.get('title_url', '')

    # Build citation based on type
    if pub_type == 'papers':
        citation = build_paper_citation(pub)
        links_html = build_links_for_paper(pub)
    elif pub_type == 'chapters':
        citation = build_chapter_citation(pub)
        links_html = build_links_for_chapter(pub)
    elif pub_type == 'dissertations':
        citation = build_dissertation_citation(pub)
        links_html = build_links_for_dissertation(pub)
    elif pub_type == 'talks':
        citation = build_talk_citation(pub)
        links_html = build_links_for_talk(pub)
    elif pub_type == 'courses':
        citation = build_course_citation(pub)
        links_html = build_links_for_course(pub)
    elif pub_type == 'posters':
        citation = build_poster_citation(pub)
        links_html = build_links_for_poster(pub)
    else:
        citation = ''
        links_html = ''

    # Build image path - prepend directory if not empty
    if image:
        image_src = f"images/publications/{image}"
    else:
        image_src = ""

    # Build title HTML
    if title_url:
        # Resolve the URL (handles local file paths)
        resolved_url = resolve_link(title_url)
        title_html = f'<a href="{resolved_url}" target="_blank">{title}</a>'
    else:
        title_html = title

    # Build links paragraph if we have links
    links_p = ''
    if links_html:
        links_p = f'\n                        <p class="publication-links">{links_html}</p>'

    # Build citation paragraph if we have citation
    citation_p = ''
    if citation:
        citation_p = f'\n                        <p>{citation}</p>'

    html = f'''                <div class="publication-card">
                    <img src="{image_src}" alt="Publication thumbnail">
                    <div>
                        <h4>{title_html}</h4>{citation_p}{links_p}
                    </div>
                </div>'''

    return html


def generate_section_content(publications: List[Dict[str, Any]], pub_type: str) -> str:
    """Generate HTML content for a publications section.

    Args:
        publications: List of publication dictionaries
        pub_type: Type of publication

    Returns:
        HTML string with all publication cards for the section
    """
    cards = [generate_publication_card(pub, pub_type) for pub in publications]
    return '\n\n'.join(cards)


def build_publications(
    data_path: Path,
    template_path: Path,
    output_path: Path
) -> None:
    """Build publications.html from data and template.

    Args:
        data_path: Path to publications.xlsx
        template_path: Path to template HTML file
        output_path: Path for generated HTML file
    """
    # Load data
    data = load_publications(data_path)

    # Generate content for each section
    replacements = {
        'PAPERS_CONTENT': generate_section_content(data.get('papers', []), 'papers'),
        'CHAPTERS_CONTENT': generate_section_content(data.get('chapters', []), 'chapters'),
        'DISSERTATIONS_CONTENT': generate_section_content(data.get('dissertations', []), 'dissertations'),
        'TALKS_CONTENT': generate_section_content(data.get('talks', []), 'talks'),
        'COURSES_CONTENT': generate_section_content(data.get('courses', []), 'courses'),
        'POSTERS_CONTENT': generate_section_content(data.get('posters', []), 'posters'),
    }

    # Inject into template
    inject_content(template_path, output_path, replacements)

    # Report
    total = sum(len(items) for items in data.values())
    print(f"Generated {output_path} with {total} publications")


def main():
    """Main entry point for CLI usage."""
    project_root = Path(__file__).parent.parent
    data_path = project_root / 'data' / 'publications.xlsx'
    template_path = project_root / 'templates' / 'publications.html'
    output_path = project_root / 'publications.html'

    build_publications(data_path, template_path, output_path)


if __name__ == '__main__':
    main()
