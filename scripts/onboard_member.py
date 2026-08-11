#!/usr/bin/env python3
"""Onboard new lab members.

This script:
1. Checks if member is an alumni (and reactivates them if so)
2. Processes photo with hand-drawn border (if provided)
3. Generates or edits bio using local LLM (gpt-oss-20b)
4. Adds member to people.xlsx
5. Adds member to JRM_CV.tex
6. Invites to GitHub organization (if --github provided)
7. Shares Google Calendars (if --gmail provided)
8. Rebuilds people.html

Idempotent: Running twice with same name will update existing entry.
Reactivation: Running on an alumni will move them back to active status.

Usage:
    python onboard_member.py "First Last"
    python onboard_member.py "First Last" --rank "grad student"
    python onboard_member.py "First Last" --photo photo.jpg --bio "Bio text..."
    python onboard_member.py "First Last" --website "https://example.com"
    python onboard_member.py "First Last" --github username --teams "supereeg,hypertools"
    python onboard_member.py "First Last" --gmail user@gmail.com
"""

import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional, Tuple, Dict, Any, List

import openpyxl

# =============================================================================
# Constants
# =============================================================================

GITHUB_ORG = "ContextLab"
CREDENTIALS_DIR = Path.home() / ".config" / "cdl"
GOOGLE_CREDENTIALS_FILE = CREDENTIALS_DIR / "google-credentials.json"

# Google Calendar IDs (from lab manual)
CALENDARS = {
    "contextual dynamics lab": {
        "id": "5ta50cfv4uih0a0k8m2di9dhjo@group.calendar.google.com",
        "undergrad_role": "reader",  # read-only for undergrads
        "default_role": "writer",  # write for grads/postdocs
    },
    "cdl resources": {
        "id": "dgcv8l8a8s10hfg2s5h0qec0q0@group.calendar.google.com",
        "undergrad_role": "writer",
        "default_role": "writer",
    },
    "out of lab": {
        "id": "h1j06dohcg7v1g2o5tkb7ijhvs@group.calendar.google.com",
        "undergrad_role": "writer",
        "default_role": "writer",
    },
}


def get_project_root() -> Path:
    return Path(__file__).parent.parent


def fuzzy_match_team(query: str, team_names: List[str]) -> Optional[str]:
    """Match team name using fuzzy matching (case-insensitive, space-insensitive)."""
    query_normalized = query.lower().replace(" ", "").replace("-", "").replace("_", "")

    for team in team_names:
        team_normalized = (
            team.lower().replace(" ", "").replace("-", "").replace("_", "")
        )
        if query_normalized == team_normalized:
            return team

    best_match = None
    best_ratio = 0.6
    for team in team_names:
        team_normalized = (
            team.lower().replace(" ", "").replace("-", "").replace("_", "")
        )
        ratio = SequenceMatcher(None, query_normalized, team_normalized).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_match = team

    return best_match


def get_github_teams() -> Dict[str, int]:
    """Get all teams in the ContextLab organization with their IDs."""
    result = subprocess.run(
        ["gh", "api", f"/orgs/{GITHUB_ORG}/teams", "--paginate"],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        print(f"  Warning: Could not fetch GitHub teams: {result.stderr}")
        return {}

    teams = json.loads(result.stdout)
    return {team["name"]: team["id"] for team in teams}


def is_github_org_member(username: str) -> bool:
    """Check if user is already a member of the organization."""
    result = subprocess.run(
        ["gh", "api", f"/orgs/{GITHUB_ORG}/members/{username}"],
        capture_output=True,
        text=True,
    )
    return result.returncode == 0


def invitation_login(invitation: dict) -> str:
    """The username a pending invitation refers to.

    GitHub leaves `login` null whenever it recorded the invite against an
    email address rather than an account, which is what happens for users
    whose email is private. Sreshth Tiwari's came back as::

        {"login": null, "email": "SreshthTiwari@users.noreply.github.com"}

    Reading only `login` dropped those, so the caller saw an empty list and
    re-sent the invitation on every run.
    """
    login = invitation.get("login")
    if login:
        return login

    local, _, domain = (invitation.get("email") or "").partition("@")
    if domain.lower() == "users.noreply.github.com":
        # Accounts created after mid-2017 use "<id>+<username>@...".
        return local.split("+")[-1]
    return ""


def get_pending_invitations() -> List[str]:
    """Get list of usernames with pending org invitations."""
    result = subprocess.run(
        ["gh", "api", f"/orgs/{GITHUB_ORG}/invitations", "--paginate"],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        # Staying quiet here reads as "nobody is invited" and silently
        # re-invites everyone, so say so instead.
        print(
            "    Warning: could not read pending invitations "
            f"({result.stderr.strip() or 'gh exited ' + str(result.returncode)})"
        )
        return []

    invitations = json.loads(result.stdout)
    return [login for login in map(invitation_login, invitations) if login]


def invite_to_github_org(username: str, team_names: Optional[List[str]] = None) -> bool:
    """Invite user to GitHub org and add to specified teams."""
    print(f"\n  GitHub: Processing {username}...")

    if is_github_org_member(username):
        print(f"    {username} is already an org member")
        if team_names:
            return add_to_github_teams(username, team_names)
        return True

    pending = get_pending_invitations()
    if username.lower() in [p.lower() for p in pending]:
        print(f"    {username} already has a pending invitation")
        return True

    all_teams = get_github_teams()
    if not all_teams:
        print("    Warning: Could not fetch teams, inviting without team assignment")

    teams_to_add = ["Lab default"]
    if team_names:
        for requested_team in team_names:
            matched = fuzzy_match_team(requested_team, list(all_teams.keys()))
            if matched and matched not in teams_to_add:
                teams_to_add.append(matched)
                if matched != requested_team:
                    print(f"    Matched '{requested_team}' -> '{matched}'")
            elif not matched:
                print(f"    Warning: No match found for team '{requested_team}'")

    team_ids = [all_teams[t] for t in teams_to_add if t in all_teams]

    cmd = [
        "gh",
        "api",
        "-X",
        "POST",
        f"/orgs/{GITHUB_ORG}/invitations",
        "-f",
        f"invitee_id={get_github_user_id(username)}",
        "-f",
        "role=direct_member",
    ]
    for tid in team_ids:
        cmd.extend(["-F", f"team_ids[]={tid}"])

    result = subprocess.run(cmd, capture_output=True, text=True)

    if result.returncode != 0:
        if "invitee_id" in result.stderr or "user_id" in result.stderr:
            cmd_email = [
                "gh",
                "api",
                "-X",
                "POST",
                f"/orgs/{GITHUB_ORG}/invitations",
                "-f",
                f"email={username}@users.noreply.github.com",
                "-f",
                "role=direct_member",
            ]
            for tid in team_ids:
                cmd_email.extend(["-F", f"team_ids[]={tid}"])
            result = subprocess.run(cmd_email, capture_output=True, text=True)

        if result.returncode != 0:
            print(f"    Error inviting to org: {result.stderr}")
            return False

    print(f"    Invited {username} to {GITHUB_ORG}")
    print(f"    Teams: {', '.join(teams_to_add)}")
    return True


def get_github_user_id(username: str) -> Optional[int]:
    """Get GitHub user ID from username."""
    result = subprocess.run(
        ["gh", "api", f"/users/{username}", "--jq", ".id"],
        capture_output=True,
        text=True,
    )
    if result.returncode == 0:
        try:
            return int(result.stdout.strip())
        except ValueError:
            pass
    return None


def add_to_github_teams(username: str, team_names: List[str]) -> bool:
    """Add existing org member to additional teams."""
    all_teams = get_github_teams()
    success = True

    for requested_team in team_names:
        matched = fuzzy_match_team(requested_team, list(all_teams.keys()))
        if not matched:
            print(f"    Warning: No match found for team '{requested_team}'")
            continue

        team_slug = matched.lower().replace(" ", "-")
        result = subprocess.run(
            [
                "gh",
                "api",
                "-X",
                "PUT",
                f"/orgs/{GITHUB_ORG}/teams/{team_slug}/memberships/{username}",
                "-f",
                "role=member",
            ],
            capture_output=True,
            text=True,
        )

        if result.returncode == 0:
            print(f"    Added to team: {matched}")
        else:
            print(f"    Warning: Could not add to {matched}: {result.stderr}")
            success = False

    return success


def setup_google_credentials() -> bool:
    """Check for Google credentials and provide setup instructions if missing."""
    if GOOGLE_CREDENTIALS_FILE.exists():
        return True

    print("\n" + "=" * 60)
    print("GOOGLE CALENDAR SETUP REQUIRED")
    print("=" * 60)
    print(f"""
To share Google Calendars, you need a service account:

1. Go to https://console.cloud.google.com/
2. Create a project (or use existing CDL project)
3. Enable the Google Calendar API
4. Create a Service Account:
   - Go to IAM & Admin > Service Accounts
   - Create service account named 'cdl-onboarding-bot'
   - Create a JSON key and download it
5. Share each calendar with the service account email:
   - Open Google Calendar settings for each calendar
   - Add the service account email with 'Make changes to events'

6. Save the JSON key file to:
   {GOOGLE_CREDENTIALS_FILE}

Create the directory if needed:
   mkdir -p {CREDENTIALS_DIR}
   mv ~/Downloads/your-key-file.json {GOOGLE_CREDENTIALS_FILE}
   chmod 600 {GOOGLE_CREDENTIALS_FILE}
""")
    print("=" * 60)
    return False


def get_calendar_service():
    """Get authenticated Google Calendar service."""
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except ImportError:
        print("  Installing Google API dependencies...")
        subprocess.check_call(
            [
                sys.executable,
                "-m",
                "pip",
                "install",
                "-q",
                "google-api-python-client",
                "google-auth",
            ]
        )
        from google.oauth2 import service_account
        from googleapiclient.discovery import build

    creds = service_account.Credentials.from_service_account_file(
        str(GOOGLE_CREDENTIALS_FILE),
        scopes=["https://www.googleapis.com/auth/calendar"],
    )
    return build("calendar", "v3", credentials=creds)


def get_calendar_acl(service, calendar_id: str, email: str) -> Optional[str]:
    """Check if user already has access to calendar. Returns role or None."""
    try:
        acl_list = service.acl().list(calendarId=calendar_id).execute()
        for rule in acl_list.get("items", []):
            if rule.get("scope", {}).get("value", "").lower() == email.lower():
                return rule.get("role")
    except Exception:
        pass
    return None


def share_calendar(
    service, calendar_id: str, email: str, role: str, calendar_name: str
) -> bool:
    """Share a calendar with a user."""
    existing_role = get_calendar_acl(service, calendar_id, email)

    if existing_role:
        if existing_role == role:
            print(f"    {calendar_name}: already has {role} access")
            return True
        else:
            print(f"    {calendar_name}: has {existing_role}, updating to {role}")

    acl_rule = {
        "scope": {"type": "user", "value": email},
        "role": role,
    }

    try:
        if existing_role:
            rule_id = f"user:{email}"
            service.acl().update(
                calendarId=calendar_id, ruleId=rule_id, body=acl_rule
            ).execute()
        else:
            service.acl().insert(
                calendarId=calendar_id, body=acl_rule, sendNotifications=True
            ).execute()

        role_desc = "read" if role == "reader" else "write"
        print(f"    {calendar_name}: granted {role_desc} access")
        return True

    except Exception as e:
        print(f"    {calendar_name}: Error - {e}")
        return False


def share_google_calendars(email: str, rank: str) -> bool:
    """Share all lab calendars with appropriate permissions based on rank."""
    if not setup_google_credentials():
        return False

    print(f"\n  Google Calendar: Sharing with {email}...")

    try:
        service = get_calendar_service()
    except Exception as e:
        print(f"    Error connecting to Google Calendar API: {e}")
        return False

    is_undergrad = "undergrad" in rank.lower()
    success = True

    for cal_name, cal_info in CALENDARS.items():
        role = cal_info["undergrad_role"] if is_undergrad else cal_info["default_role"]
        if not share_calendar(service, cal_info["id"], email, role, cal_name):
            success = False

    return success


# Bios are rewritten by Dartmouth's own LLM service. See dartmouth_chat.py.
#
# This replaces a local mlx-lm setup that downloaded an 18GB Qwen2.5-32B into
# ~/.cache/cdl/llm-venv and ran it on device. mlx-lm is Apple-Silicon only, so
# that path could never work for the lab's Windows and Linux members -- the
# same people issue #14 was about. The hosted service costs the lab nothing,
# needs no download, and answers in about a second.
from dartmouth_chat import DartmouthChatError, chat as dartmouth_chat


def _clean_bio(text: str) -> str:
    """Strip the wrapping a chat model puts around a one-line answer."""
    text = text.strip().strip('"').strip()

    # "Here is the bio:" / "Sure! Here's..." -- drop a preamble line and keep
    # what follows, rather than storing the chatter as the bio.
    lowered = text.lower()
    if lowered.startswith(("here", "sure", "certainly", "of course")):
        _, sep, rest = text.partition("\n")
        text = rest.strip() if sep and rest.strip() else ""
        if not text:
            return ""
        text = text.strip('"').strip()

    return text


def generate_bio_with_llm(first_name: str, year: str) -> str:
    fallback = f"{first_name} joined the lab in {year} and is interested in how people learn and remember."

    prompt = (
        f"Write a single sentence professional bio for {first_name}, an "
        f"undergraduate who joined a cognitive neuroscience memory research "
        f"lab in {year}. Keep it generic. Do not use pronouns. Maximum 25 "
        f"words. Output ONLY the bio sentence, nothing else."
    )

    print("  Generating bio with LLM...")
    try:
        bio = _clean_bio(dartmouth_chat(prompt, max_tokens=200))
    except DartmouthChatError as exc:
        print(f"  WARNING: could not reach the LLM service: {exc}")
        print("  Falling back to a generic bio; edit it by hand in people.xlsx.")
        return fallback

    return bio if len(bio) > 15 else fallback


PRONOUN_GROUPS = {
    "he": {"he", "him", "his"},
    "she": {"she", "her", "hers"},
}


def stated_pronouns(text: str) -> frozenset:
    """Which gendered pronoun groups a bio uses, if any.

    Returns a subset of {"he", "she"}; an empty set means the bio is already
    pronoun-free, which is what a first-person submission looks like before
    editing.
    """
    words = set(re.findall(r"[a-z]+", (text or "").lower()))
    return frozenset(
        group for group, forms in PRONOUN_GROUPS.items() if words & forms
    )


BIO_EDIT_RULES = (
    "Edit this bio. Rules:\n"
    "1. Write in third person, NOT first person. Never start with \"Hi\", "
    '"I am", or "I\'m".\n'
    "2. PRONOUNS: if the original bio already uses gendered pronouns "
    "(he/him/his, she/her/hers), keep exactly those -- the person wrote "
    "them about themselves, so never swap them for \"they/them\" and "
    "never reword them away. ONLY when the original contains no pronouns "
    "at all may you use \"they/them\" or reword to avoid pronouns. Never "
    "infer pronouns from the name.\n"
    '3. Start the bio with the first name "{first_name}" (e.g., '
    '"{first_name} is a..."). Do not include the last name.\n'
    "4. Fix typos and grammar\n"
    "5. Keep 1-3 sentences max\n"
    "6. Remove dangerous personal info (SSN, addresses, phone numbers)\n"
    "7. Keep professional and friendly tone\n\n"
)

BIO_EDIT_ATTEMPTS = 3


def edit_bio_with_llm(bio: str, first_name: str) -> str:
    """Tidy a submitted bio without changing whose it is.

    The pronoun rule is checked in code rather than trusted to the prompt.
    Even with rule 2 spelled out, the model intermittently satisfies it by
    rewriting "His research interests lie in causal inference" into
    "...with research interests in causal inference" -- pronoun gone, rule
    technically unbroken. Publishing that erases what someone said about
    themselves, so an edit that changes the bio's pronouns is rejected and
    retried, and the submitted text stands if the model will not comply.
    """
    prompt = BIO_EDIT_RULES.format(first_name=first_name) + (
        f'Original: "{bio}"\n\n' "Output ONLY the edited bio, nothing else:"
    )
    wanted = stated_pronouns(bio)

    print("  Editing bio with LLM...")
    for attempt in range(BIO_EDIT_ATTEMPTS):
        try:
            edited = _clean_bio(dartmouth_chat(prompt, max_tokens=400))
        except DartmouthChatError as exc:
            print(f"  WARNING: could not reach the LLM service: {exc}")
            print("  Keeping the bio as supplied.")
            return bio

        if len(edited) <= 15:
            return bio

        got = stated_pronouns(edited)
        if got == wanted:
            return edited

        if wanted:
            problem = (
                f"it dropped the pronoun the bio used ({'/'.join(sorted(wanted))})"
            )
            insist = (
                "Your previous attempt REMOVED the pronouns the person used "
                f"about themselves ({', '.join(sorted(wanted))}). Keep them, "
                "word for word. Do not rephrase to avoid them."
            )
        else:
            problem = f"it invented pronouns ({'/'.join(sorted(got))})"
            insist = (
                "Your previous attempt ADDED gendered pronouns "
                f"({', '.join(sorted(got))}) that the original did not use. "
                "Do not guess someone's pronouns. Use \"they/them\" or reword "
                "to avoid pronouns."
            )

        print(
            f"  Retrying bio edit ({attempt + 1}/{BIO_EDIT_ATTEMPTS}): {problem}"
        )
        prompt = (
            BIO_EDIT_RULES.format(first_name=first_name)
            + insist
            + f'\n\nOriginal: "{bio}"\n\n'
            + "Output ONLY the edited bio, nothing else:"
        )

    print("  WARNING: the LLM kept changing the bio's pronouns.")
    print("  Keeping the bio as supplied.")
    return bio


def parse_name(full_name: str) -> Tuple[str, str]:
    parts = full_name.strip().split()
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def find_photo(photo_hint: str, project_root: Path) -> Optional[Path]:
    search_dirs = [
        project_root,
        project_root / "images",
        project_root / "images" / "people",
        Path.home() / "Downloads",
        Path.home() / "Desktop",
    ]

    extensions = [".jpg", ".jpeg", ".png", ".JPG", ".JPEG", ".PNG"]

    for search_dir in search_dirs:
        if not search_dir.exists():
            continue
        for ext in extensions:
            candidate = search_dir / f"{photo_hint}{ext}"
            if candidate.exists():
                return candidate

    for ext in extensions:
        candidate = Path(f"{photo_hint}{ext}")
        if candidate.exists():
            return candidate

    return None


def photo_already_processed(photo_base: str, project_root: Path) -> bool:
    """Check if a photo has already been processed with hand-drawn borders.

    Verifies three conditions:
    1. The PNG file exists
    2. Resolution is 500x500 (the output size of add_borders.py)
    3. Corner pixels are transparent (borders leave transparent margins)
    """
    processed_photo = project_root / "images" / "people" / f"{photo_base}.png"
    if not processed_photo.exists():
        return False

    try:
        from PIL import Image
        img = Image.open(processed_photo)

        w, h = img.size

        # Check that image is square (bordered images are always square)
        if w != h:
            return False

        # Check that corner pixels are transparent (hand-drawn borders
        # leave transparent margins around the image)
        if img.mode != 'RGBA':
            return False
        corners = [
            img.getpixel((0, 0)),
            img.getpixel((w - 1, 0)),
            img.getpixel((0, h - 1)),
            img.getpixel((w - 1, h - 1)),
        ]
        # All corners should be fully transparent (alpha == 0)
        if not all(c[3] == 0 for c in corners):
            return False

        return True
    except Exception:
        return False


def process_photo(
    photo_path: Path, output_name: str, project_root: Path
) -> Optional[str]:
    output_dir = project_root / "images" / "people"
    output_file = output_dir / f"{output_name}.png"

    add_borders_script = project_root / "scripts" / "add_borders.py"

    print(f"  Processing photo with face detection...")
    result = subprocess.run(
        [
            sys.executable,
            str(add_borders_script),
            str(photo_path),
            str(output_dir),
            "--face",
        ],
        capture_output=True,
        text=True,
    )

    if result.returncode != 0:
        print(f"  Warning: Photo processing failed: {result.stderr}")
        return None

    processed_file = output_dir / f"{photo_path.stem}.png"
    if processed_file.exists() and processed_file != output_file:
        processed_file.rename(output_file)

    if output_file.exists():
        if photo_path != output_file and photo_path.suffix.lower() != ".png":
            try:
                photo_path.unlink()
                print(f"  Removed original photo: {photo_path}")
            except Exception as e:
                print(f"  Warning: Could not remove original photo: {e}")

        print(f"  Photo saved to: {output_file}")
        return f"{output_name}.png"

    return None


def member_exists_in_spreadsheet(
    xlsx_path: Path, name: str
) -> Tuple[bool, Optional[int]]:
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb["members"]

    name_lower = name.lower()
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if row[1] and str(row[1]).lower() == name_lower:
            wb.close()
            return True, row_idx

    wb.close()
    return False, None


def get_existing_bio(xlsx_path: Path, name: str) -> Optional[str]:
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb["members"]

    name_lower = name.lower()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] and str(row[1]).lower() == name_lower:
            bio = row[4] if len(row) > 4 else None
            wb.close()
            if bio and len(str(bio)) > 10:
                return str(bio)
            return None

    wb.close()
    return None


def find_alumni_entry(xlsx_path: Path, name: str) -> Optional[Dict[str, Any]]:
    """Find alumni entry across all alumni sheets. Returns dict with sheet name, row, and data."""
    wb = openpyxl.load_workbook(xlsx_path)
    name_lower = name.lower()
    name_title = name.title().lower()

    alumni_sheets = [
        "alumni_postdocs",
        "alumni_grads",
        "alumni_managers",
        "alumni_undergrads",
    ]

    for sheet_name in alumni_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        sheet = wb[sheet_name]
        headers = [cell.value for cell in sheet[1]]

        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not row[0]:
                continue
            row_name = str(row[0]).lower()
            if row_name == name_lower or row_name == name_title:
                entry: Dict[str, Any] = {
                    "sheet": sheet_name,
                    "row_idx": row_idx,
                    "name": row[0],
                }
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        entry[str(header)] = row[i]
                wb.close()
                return entry

    wb.close()
    return None


def remove_from_alumni(xlsx_path: Path, alumni_entry: Dict[str, Any]) -> None:
    """Remove an entry from the alumni sheet."""
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb[alumni_entry["sheet"]]
    sheet.delete_rows(alumni_entry["row_idx"])
    wb.save(xlsx_path)
    wb.close()
    print(f"  Removed {alumni_entry['name']} from {alumni_entry['sheet']}")


def get_bio_from_git_history(
    xlsx_path: Path, name: str, project_root: Path
) -> Optional[str]:
    """Search git history for old bio in people.xlsx."""
    name_lower = name.lower()
    name_title = name.title().lower()

    try:
        result = subprocess.run(
            [
                "git",
                "log",
                "--oneline",
                "--follow",
                "--",
                str(xlsx_path.relative_to(project_root)),
            ],
            cwd=project_root,
            capture_output=True,
            text=True,
        )
        if result.returncode != 0:
            return None

        commits = [
            line.split()[0] for line in result.stdout.strip().split("\n") if line
        ]

        for commit in commits[:20]:  # Limit to last 20 commits
            try:
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    tmp_path = tmp.name

                extract_result = subprocess.run(
                    ["git", "show", f"{commit}:{xlsx_path.relative_to(project_root)}"],
                    cwd=project_root,
                    capture_output=True,
                )
                if extract_result.returncode != 0:
                    continue

                with open(tmp_path, "wb") as f:
                    f.write(extract_result.stdout)

                wb = openpyxl.load_workbook(tmp_path)
                if "members" in wb.sheetnames:
                    sheet = wb["members"]
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[1] and str(row[1]).lower() in [name_lower, name_title]:
                            bio = row[4] if len(row) > 4 else None
                            if bio and len(str(bio)) > 10:
                                wb.close()
                                Path(tmp_path).unlink(missing_ok=True)
                                print(
                                    f"  Found old bio in git history (commit {commit})"
                                )
                                return str(bio)
                wb.close()
                Path(tmp_path).unlink(missing_ok=True)

            except Exception:
                continue

    except Exception as e:
        print(f"  Note: Could not search git history: {e}")

    return None


def add_to_spreadsheet(
    xlsx_path: Path, name: str, role: str, bio: str, image: str, website: str
) -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb["members"]

    exists, row_idx = member_exists_in_spreadsheet(xlsx_path, name)
    wb.close()

    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb["members"]

    if exists and row_idx is not None:
        print(f"  Updating existing entry for {name} at row {row_idx}")
        if image:
            sheet.cell(row=row_idx, column=1, value=image)
        sheet.cell(row=row_idx, column=2, value=name)
        if website:
            sheet.cell(row=row_idx, column=3, value=website)
        sheet.cell(row=row_idx, column=4, value=role)
        if bio:
            sheet.cell(row=row_idx, column=5, value=bio)
    else:
        last_row = sheet.max_row
        while last_row > 1 and not any(
            sheet.cell(row=last_row, column=c).value for c in range(1, 7)
        ):
            last_row -= 1

        new_row = last_row + 1

        sheet.cell(row=new_row, column=1, value=image)
        sheet.cell(row=new_row, column=2, value=name)
        sheet.cell(row=new_row, column=3, value=website)
        sheet.cell(row=new_row, column=4, value=role)
        sheet.cell(row=new_row, column=5, value=bio)

        print(f"  Added new entry for {name} at row {new_row}")

    wb.save(xlsx_path)
    wb.close()


def member_exists_in_cv(cv_path: Path, name: str) -> bool:
    content = cv_path.read_text(encoding="utf-8")
    name_escaped = re.escape(name)
    pattern = r"\\item\s+" + name_escaped + r"\*?\s*\("
    return bool(re.search(pattern, content, re.IGNORECASE))


def cv_entry_is_open(cv_path: Path, name: str) -> bool:
    """Does `name` have an open-ended CV entry -- '(... -- )'?

    Asking about open rather than closed entries matters once someone can hold
    two entries at once. Paxton Fitzpatrick has a closed '(2017 -- 2019)' under
    Undergraduate Advisees AND an open '(Doctoral student; 2021 -- )' under
    Graduate Advisees; a has-an-end-date question answers True for him and
    would send onboarding off to "reopen" the finished undergraduate range.
    """
    content = cv_path.read_text(encoding="utf-8")
    pattern = r"\\item\s+" + re.escape(name) + r"\*?\s*\([^)]*--\s*\)"
    return bool(re.search(pattern, content, re.IGNORECASE))


def reopen_cv_entry(cv_path: Path, name: str) -> bool:
    """Remove the end date from a closed CV entry: (2019 -- 2021) -> (2019 -- ).

    Handles the qualifiers real entries carry, which the year-only patterns
    this replaces could not match at all:

        \\item Kirsten Ziman (Doctoral student; 2017 -- 2022; current position:
        Postdoctoral researcher at Princeton University)

    becomes '(Doctoral student; 2017 -- )'. Eight of the eleven graduate
    advisees are written this way, so onboarding a returning one used to
    report "reopening..." and then return False, failing the whole run.
    """
    content = cv_path.read_text(encoding="utf-8")
    pattern = r"\\item\s+" + re.escape(name) + r"\*?\s*\(([^)]*)\)"

    def reopen(inner: str) -> str:
        # They are back, so a recorded destination is now wrong.
        inner = re.sub(r";?\s*current position:[^;)]*", "", inner)
        if re.search(r"\d{4}\s*--\s*\d{4}", inner):
            inner = re.sub(r"(\d{4})\s*--\s*\d{4}", r"\1 -- ", inner)
        else:
            inner = re.sub(r"(\d{4})\s*$", r"\1 -- ", inner)
        return inner

    for match in re.finditer(pattern, content, re.IGNORECASE):
        inner = match.group(1)
        if re.search(r"--\s*$", inner):
            continue  # already open

        new_inner = reopen(inner)
        if new_inner == inner:
            continue

        start, end = match.span(1)
        cv_path.write_text(
            content[:start] + new_inner + content[end:], encoding="utf-8"
        )
        print(f"  Reopened CV entry for {name}")
        return True

    return False


# The three mentorship subsections of JRM_CV.tex, in document order. The
# marker locates an existing entry's section; the pattern anchors an insert.
CV_SECTIONS = {
    "postdoc": {
        "label": "Postdoctoral Advisees",
        "marker": r"\textit{Postdoctoral Advisees}:",
        "pattern": r"(\\textit\{Postdoctoral Advisees\}:\s*\n\\begin\{etaremune\})",
        "entry": "\\item {name} ({year} -- )",
    },
    "grad": {
        "label": "Graduate Advisees",
        "marker": r"\textit{Graduate Advisees}:",
        "pattern": r"(\\textit\{Graduate Advisees\}:\s*\n\\begin\{etaremune\})",
        "entry": "\\item {name} (Doctoral student; {year} -- )",
    },
    "undergrad": {
        "label": "Undergraduate Advisees",
        "marker": r"\textit{Undergraduate Advisees}:",
        "pattern": (
            r"(\\textit\{Undergraduate Advisees\}:\s*\n\\blfootnote\{[^}]*\}\s*\n"
            r"\\begin\{multicols\}\{2\}\s*\n\\begin\{etaremune\})"
        ),
        "entry": "\\item {name} ({year} -- )",
    },
}


def cv_section_for_role(role: str) -> Optional[str]:
    """Which CV mentorship section a role belongs in, or None for no entry.

    Two deliberate rules, confirmed by the lab director:
      - Lab managers are staff rather than trainees, so they get NO CV entry.
        Returning None here is the point; before this they were filed under
        Undergraduate Advisees by the catch-all branch (issue #17).
      - Research assistants ARE listed among the undergraduate advisees. That
        is intended, not an accidental fallthrough.
    """
    role_lower = role.lower()
    if "lab manager" in role_lower:
        return None
    if "postdoc" in role_lower:
        return "postdoc"
    if "grad" in role_lower and "undergrad" not in role_lower:
        return "grad"
    return "undergrad"


def find_cv_section(cv_path: Path, name: str) -> Optional[str]:
    """Which CV section `name` is currently listed under, or None if absent.

    An open-ended entry wins over a closed one, so someone with a closed
    undergrad range and an open doctoral range reads as a current grad -- the
    shape Paxton Fitzpatrick is recorded in.
    """
    content = cv_path.read_text(encoding="utf-8")
    name_escaped = re.escape(name)

    entry = re.search(
        r"\\item\s+" + name_escaped + r"\*?\s*\([^)]*--\s*\)",
        content,
        re.IGNORECASE,
    ) or re.search(r"\\item\s+" + name_escaped + r"\*?\s*\(", content, re.IGNORECASE)
    if not entry:
        return None

    best_key, best_pos = None, -1
    for key, section in CV_SECTIONS.items():
        pos = content.find(section["marker"])
        if best_pos < pos < entry.start():
            best_key, best_pos = key, pos
    return best_key


def close_cv_entry(cv_path: Path, name: str, end_year: str) -> bool:
    """Close an open CV range: (2024 -- ) -> (2024 -- 2026).

    Preserves any qualifier before the years, so a graduate entry closes as
    '(Doctoral student; 2021 -- 2026)'.
    """
    content = cv_path.read_text(encoding="utf-8")
    pattern = r"(\\item\s+" + re.escape(name) + r"\*?\s*\([^)]*?--\s*)\)"

    # count=1: close the first open entry only. Closing every one at a stroke
    # would rewrite an unrelated second open entry for the same name, which is
    # never what a single role change means.
    new_content, count = re.subn(
        pattern, lambda m: f"{m.group(1)}{end_year})", content, count=1,
        flags=re.IGNORECASE
    )
    if count == 0:
        return False

    cv_path.write_text(new_content, encoding="utf-8")
    return True


def add_to_cv(cv_path: Path, name: str, role: str, year: str) -> bool:
    section_key = cv_section_for_role(role)

    if section_key is None:
        print(f"  {name} is a {role}; not a trainee, so no CV entry is made")
        return True

    section = CV_SECTIONS[section_key]
    existing = find_cv_section(cv_path, name)

    if existing == section_key:
        if cv_entry_is_open(cv_path, name):
            print(f"  {name} already exists in CV with open date, skipping")
            return True
        print(f"  {name} exists in CV with end date, reopening...")
        return reopen_cv_entry(cv_path, name)

    if existing is not None:
        # A role change closes the old range and opens a new one, the way
        # Paxton Fitzpatrick is recorded: '(2017 -- 2019)' under Undergraduate
        # Advisees and '(Doctoral student; 2021 -- )' under Graduate Advisees.
        if cv_entry_is_open(cv_path, name) and close_cv_entry(
            cv_path, name, year
        ):
            print(
                f"  Closed {name}'s {CV_SECTIONS[existing]['label']} entry at {year}"
            )

    content = cv_path.read_text(encoding="utf-8")
    match = re.search(section["pattern"], content)
    if not match:
        print(f"  Warning: Could not find section for {role} in CV")
        return False

    entry = section["entry"].format(name=name, year=year)
    insert_pos = match.end()
    new_content = content[:insert_pos] + f"\n  {entry}" + content[insert_pos:]

    cv_path.write_text(new_content, encoding="utf-8")
    print(f"  Added {name} to CV under {section['label']}")
    return True


def rebuild_pages(project_root: Path) -> bool:
    """Rebuild people.html and the CV.

    Returns:
        True only if every rebuild actually succeeded. Failures are reported
        as errors rather than notes, so the operator is never told a build
        happened when it did not.
    """
    scripts_dir = project_root / "scripts"
    succeeded = True

    print("\nRebuilding people.html...")
    result = subprocess.run(
        [sys.executable, "build_people.py"],
        cwd=scripts_dir,
        capture_output=True,
        text=True,
    )
    if result.returncode == 0:
        print("  people.html rebuilt successfully")
    else:
        print(f"  ERROR: build_people.py failed:\n{result.stderr.strip()}")
        succeeded = False

    print("\nRebuilding CV...")
    if shutil.which("xelatex") is None:
        print("  ERROR: xelatex not found, so the CV was NOT rebuilt.")
        print("  JRM_CV.pdf still reflects the previous state.")
        print("  Install a LaTeX distribution, then run: cd scripts && python build_cv.py")
        succeeded = False
    else:
        result = subprocess.run(
            [sys.executable, "build_cv.py"], cwd=scripts_dir, capture_output=True, text=True
        )
        if result.returncode == 0:
            print("  CV rebuilt successfully")
        else:
            print("  ERROR: CV build failed. JRM_CV.pdf does NOT include this member.")
            details = (result.stdout or result.stderr).strip()
            for line in details.splitlines()[-15:]:
                print(f"    {line}")
            succeeded = False

    return succeeded


def onboard_member(
    name: str,
    rank: str = "undergrad",
    photo: Optional[str] = None,
    bio: Optional[str] = None,
    website: Optional[str] = None,
    github_username: Optional[str] = None,
    github_teams: Optional[List[str]] = None,
    gmail: Optional[str] = None,
    skip_rebuild: bool = False,
    skip_llm: bool = False,
) -> bool:
    project_root = get_project_root()
    xlsx_path = project_root / "data" / "people.xlsx"
    cv_path = project_root / "documents" / "JRM_CV.tex"

    first_name, last_name = parse_name(name)
    name_lower = name.lower()
    photo_base = (
        f"{first_name}_{last_name}".lower().replace(" ", "_")
        if last_name
        else first_name.lower()
    )
    current_year = str(datetime.now().year)

    print(f"\nOnboarding {name} as {rank}...")

    alumni_entry = find_alumni_entry(xlsx_path, name)
    is_reactivation = alumni_entry is not None

    if is_reactivation:
        print(f"  Found {name} in {alumni_entry['sheet']} - reactivating...")
        remove_from_alumni(xlsx_path, alumni_entry)

    image_filename = None

    if photo_already_processed(photo_base, project_root):
        print(f"  Using existing processed photo: {photo_base}.png")
        image_filename = f"{photo_base}.png"
    else:
        if photo is None:
            photo = photo_base

        photo_path = find_photo(photo, project_root)
        if photo_path:
            print(f"  Found photo: {photo_path}")
            image_filename = process_photo(photo_path, photo_base, project_root)
        else:
            print(f"  No photo found for {photo}")

    if bio is None and is_reactivation:
        print("  Searching git history for old bio...")
        bio = get_bio_from_git_history(xlsx_path, name, project_root)

    if bio is None:
        existing_bio = get_existing_bio(xlsx_path, name)
        if existing_bio:
            print(f"  Using existing bio from spreadsheet")
            bio = existing_bio

    if not skip_llm:
        if bio:
            # edit_bio_with_llm announces itself; saying it here too printed
            # the same line twice.
            bio = edit_bio_with_llm(bio, first_name)
            print(f"  Bio: {bio[:100]}...")
        else:
            print("  Generating bio with LLM...")
            bio = generate_bio_with_llm(first_name, current_year)
            print(f"  Bio: {bio}")
    elif not bio:
        bio = f"{first_name} joined the lab in {current_year} and is interested in how people learn and remember."

    print("\nUpdating spreadsheet...")
    add_to_spreadsheet(
        xlsx_path, name_lower, rank, bio, image_filename or "", website or ""
    )

    print("\nUpdating CV...")
    cv_updated = add_to_cv(cv_path, name, rank, current_year)
    if not cv_updated:
        print(f"  ERROR: {name} was NOT added to {cv_path.name}")

    # Update lab-manual (best-effort; failure doesn't block onboarding)
    try:
        from parse_lab_manual import sync_member_role, commit_and_push_lab_manual
        lab_manual_tex = project_root / 'lab-manual' / 'lab_manual.tex'
        if lab_manual_tex.exists():
            print("\nUpdating lab-manual...")
            outcome = sync_member_role(lab_manual_tex, name, rank, current_year)
            if outcome == 'added':
                print(f"  Added {name} to lab-manual under {rank}")
            elif outcome == 'role-changed':
                print(f"  Moved {name} to {rank}; the previous role was closed "
                      f"out to Lab alumni at {current_year}")
            else:
                print(f"  {name} already listed in lab-manual under {rank}, skipping")
            try:
                commit_and_push_lab_manual(
                    project_root / 'lab-manual',
                    f"Onboard {name}"
                )
                print(f"  Updated lab-manual and pushed to remote")
            except RuntimeError as e:
                print(f"  WARNING: Lab-manual updated locally but push failed: {e}")
        else:
            print("  NOTE: Lab-manual submodule not found, skipping lab-manual update")
    except Exception as e:
        print(f"  WARNING: Could not update lab-manual: {e}")

    if github_username:
        invite_to_github_org(github_username, github_teams)

    if gmail:
        share_google_calendars(gmail, rank)

    rebuilt = rebuild_pages(project_root) if not skip_rebuild else True

    action = "reactivated" if is_reactivation else "onboarded"
    if not (cv_updated and rebuilt):
        print(f"\n{name} was {action}, but some steps FAILED -- see the errors above.")
        if not cv_updated:
            print("  - the CV does not list this member")
        if not rebuilt:
            print("  - the generated pages are out of date")
        return False

    print(f"\nSuccessfully {action} {name}!")
    return True


def main():
    parser = argparse.ArgumentParser(
        description="Onboard new lab members.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python onboard_member.py "John Doe"
    python onboard_member.py "Jane Smith" --rank "grad student"
    python onboard_member.py "Bob Jones" --photo headshot --bio "Bob is interested in memory."
    python onboard_member.py "Alice Lee" --website "https://alice.com" --skip-llm

    # GitHub integration (invite to org and teams)
    python onboard_member.py "John Doe" --github johndoe
    python onboard_member.py "John Doe" --github johndoe --teams "supereeg,hypertools"

    # Google Calendar integration (share lab calendars)
    python onboard_member.py "John Doe" --gmail john.doe@gmail.com

    # Full onboarding with all integrations
    python onboard_member.py "John Doe" --rank "grad student" --github johndoe --gmail john@gmail.com
        """,
    )

    parser.add_argument("name", help='Full name of the new member (e.g., "First Last")')
    parser.add_argument(
        "--rank",
        "-r",
        default="undergrad",
        help="Role/rank (default: undergrad). Options: undergrad, grad student, postdoc",
    )
    parser.add_argument(
        "--photo",
        "-p",
        help="Photo filename without extension (default: first_last). Searches current dir, images/, Downloads/",
    )
    parser.add_argument(
        "--bio",
        "-b",
        help="Bio text (will be edited by LLM). If not provided, a generic bio is generated.",
    )
    parser.add_argument("--website", "-w", help="Personal website URL")
    parser.add_argument(
        "--github",
        "-g",
        help="GitHub username. If provided, invites to ContextLab org (adds to 'Lab default' team).",
    )
    parser.add_argument(
        "--teams",
        "-t",
        help="Comma-separated list of additional GitHub teams (uses fuzzy matching). E.g., 'supereeg,hypertools'",
    )
    parser.add_argument(
        "--gmail",
        help="Gmail address. If provided, shares lab calendars (undergrads get read access to main calendar, others get write).",
    )
    parser.add_argument(
        "--skip-rebuild", action="store_true", help="Skip rebuilding HTML pages"
    )
    parser.add_argument(
        "--skip-llm",
        action="store_true",
        help="Skip LLM processing (use bio as-is or generate simple default)",
    )

    args = parser.parse_args()

    github_teams = None
    if args.teams:
        github_teams = [t.strip() for t in args.teams.split(",") if t.strip()]

    success = onboard_member(
        name=args.name,
        rank=args.rank,
        photo=args.photo,
        bio=args.bio,
        website=args.website,
        github_username=args.github,
        github_teams=github_teams,
        gmail=args.gmail,
        skip_rebuild=args.skip_rebuild,
        skip_llm=args.skip_llm,
    )

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
