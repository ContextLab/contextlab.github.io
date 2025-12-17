#!/usr/bin/env python3
"""Build software.html from spreadsheet data.

Reads data from data/software.xlsx and generates software.html
using the template in templates/software.html.

The spreadsheet uses structured fields for each software item:
- name: Software name
- description: Software description (supports markdown/HTML)
- github_link: GitHub repository URL
- pypi_link: PyPI package URL
- docs_link: Documentation URL
- fileexchange_link: MATLAB File Exchange URL
- extra_links: Additional links in format "Label1:URL1;Label2:URL2"
"""
from pathlib import Path
from typing import List, Dict, Any
import openpyxl

from utils import inject_content
from citation_utils import markdown_to_html, resolve_link, build_links_html


def load_software(xlsx_path: Path) -> Dict[str, List[Dict[str, Any]]]:
    """Load all software data from Excel spreadsheet.

    Args:
        xlsx_path: Path to the software.xlsx file

    Returns:
        Dictionary with keys for each section (python, javascript, matlab)
        Each value is a list of software item dictionaries.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    data = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not any(cell is not None for cell in row):
                continue

            row_dict = {}
            for header, value in zip(headers, row):
                if value is None:
                    row_dict[header] = ''
                else:
                    row_dict[header] = str(value) if not isinstance(value, str) else value
            rows.append(row_dict)

        data[sheet_name] = rows

    wb.close()
    return data


def parse_extra_links(extra_links: str) -> List[tuple]:
    """Parse extra_links field into list of (label, url) tuples.

    Format: "Label1:URL1;Label2:URL2"
    """
    if not extra_links:
        return []

    links = []
    for part in extra_links.split(';'):
        if ':' in part:
            # Split on first colon only (URLs contain colons)
            label, url = part.split(':', 1)
            links.append((label.strip(), url.strip()))
    return links


def build_links_for_software(item: Dict[str, Any]) -> str:
    """Build links HTML for a software item."""
    links = []

    if item.get('github_link'):
        links.append(('GitHub', item['github_link']))
    if item.get('pypi_link'):
        links.append(('PyPI', item['pypi_link']))
    if item.get('docs_link'):
        links.append(('Docs', item['docs_link']))
    if item.get('fileexchange_link'):
        links.append(('MATLAB Central File Exchange', item['fileexchange_link']))

    # Add extra links
    links.extend(parse_extra_links(item.get('extra_links', '')))

    return build_links_html(links)


def generate_software_item(item: Dict[str, Any]) -> str:
    """Generate HTML for a single software item.

    Args:
        item: Dictionary with software data

    Returns:
        HTML string for the software item paragraph
    """
    name = item.get('name', '')
    description = item.get('description', '')

    # Convert markdown to HTML in description
    description = markdown_to_html(description)

    # Build links HTML
    links_html = build_links_for_software(item)

    # Build the paragraph
    parts = []
    if name:
        parts.append(f'<strong>{name}.</strong>')
    if description:
        parts.append(description)
    if links_html:
        parts.append(links_html)

    return f'<p>{" ".join(parts)}</p>'


def generate_section_content(items: List[Dict[str, Any]]) -> str:
    """Generate HTML content for a software section.

    Args:
        items: List of software item dictionaries

    Returns:
        HTML string with all software items for the section
    """
    if not items:
        return ''

    paragraphs = [generate_software_item(item) for item in items]
    return '\n\n                '.join(paragraphs)


def build_software(
    data_path: Path,
    template_path: Path,
    output_path: Path
) -> None:
    """Build software.html from data and template.

    Args:
        data_path: Path to software.xlsx
        template_path: Path to template HTML file
        output_path: Path for generated HTML file
    """
    # Load data
    data = load_software(data_path)

    # Generate content for each section
    replacements = {
        'PYTHON_CONTENT': generate_section_content(data.get('python', [])),
        'JAVASCRIPT_CONTENT': generate_section_content(data.get('javascript', [])),
        'MATLAB_CONTENT': generate_section_content(data.get('matlab', [])),
    }

    # Inject into template
    inject_content(template_path, output_path, replacements)

    # Report
    total = sum(len(items) for items in data.values())
    print(f"Generated {output_path} with {total} software items")


def main():
    """Main entry point for CLI usage."""
    project_root = Path(__file__).parent.parent
    data_path = project_root / 'data' / 'software.xlsx'
    template_path = project_root / 'templates' / 'software.html'
    output_path = project_root / 'software.html'

    build_software(data_path, template_path, output_path)


if __name__ == '__main__':
    main()
