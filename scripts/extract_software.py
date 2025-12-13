#!/usr/bin/env python3
"""Extract software data from existing HTML into Excel spreadsheet.

This is a one-time script to migrate existing HTML data to the spreadsheet format.
"""
import re
from pathlib import Path
from bs4 import BeautifulSoup
import openpyxl


def extract_software(html_path: Path) -> dict:
    """Extract all software data from HTML file.

    Returns dict with keys: python, javascript, matlab
    """
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f.read(), 'html.parser')

    sections = {
        'python': 'python',
        'javascript': 'javascript',
        'matlab': 'matlab'
    }

    data = {}

    for key, section_id in sections.items():
        section = soup.find('section', id=section_id)
        if not section:
            print(f"Warning: Section '{section_id}' not found")
            data[key] = []
            continue

        software_list = section.find('div', class_='software-list')
        if not software_list:
            print(f"Warning: No software-list in '{section_id}'")
            data[key] = []
            continue

        items = []
        for p in software_list.find_all('p', recursive=False):
            item = extract_software_item(p)
            if item['name']:
                items.append(item)

        data[key] = items
        print(f"Extracted {len(items)} items from '{key}'")

    return data


def extract_software_item(p) -> dict:
    """Extract data from a single software paragraph.

    Format: <strong>Name.</strong> Description. [<a>Link</a>]
    """
    item = {
        'name': '',
        'description': '',
        'links_html': ''
    }

    # Get the name from <strong> tag
    strong = p.find('strong')
    if strong:
        name = strong.get_text(strip=True)
        # Remove trailing period if present
        item['name'] = name.rstrip('.')

    # Get the full inner HTML
    inner_html = get_inner_html(p)

    # Extract links at the end (text in square brackets)
    # Links are usually at the end like [<a href="...">GitHub</a>]
    link_match = re.search(r'\[([^\]]*<a[^>]*>[^<]*</a>[^\]]*)\]', inner_html)
    if link_match:
        item['links_html'] = f'[{link_match.group(1)}]'
        # Remove the link from the HTML to get description
        inner_html = inner_html[:link_match.start()].strip()

    # Now extract description (everything after the name)
    # Remove the <strong>...</strong> part
    desc_html = re.sub(r'<strong>[^<]*</strong>\s*', '', inner_html)
    # Clean up the description
    desc_html = desc_html.strip()
    # Remove leading period if present
    if desc_html.startswith('.'):
        desc_html = desc_html[1:].strip()

    item['description'] = desc_html

    return item


def get_inner_html(element) -> str:
    """Get the inner HTML of an element as a string."""
    return ''.join(str(child) for child in element.children).strip()


def save_to_excel(data: dict, output_path: Path):
    """Save extracted data to Excel spreadsheet with multiple sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_configs = [
        ('python', ['name', 'description', 'links_html']),
        ('javascript', ['name', 'description', 'links_html']),
        ('matlab', ['name', 'description', 'links_html']),
    ]

    for sheet_name, columns in sheet_configs:
        ws = wb.create_sheet(title=sheet_name)

        # Write headers
        for col, header in enumerate(columns, 1):
            ws.cell(row=1, column=col, value=header)

        # Write data
        items = data.get(sheet_name, [])
        for row_num, item in enumerate(items, 2):
            for col, header in enumerate(columns, 1):
                value = item.get(header, '')
                ws.cell(row=row_num, column=col, value=value)

        # Adjust column widths
        for col in range(1, len(columns) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 60

    wb.save(output_path)
    print(f"Saved to {output_path}")


def main():
    project_root = Path(__file__).parent.parent
    html_path = project_root / 'software.html'
    output_path = project_root / 'data' / 'software.xlsx'

    print(f"Extracting from: {html_path}")
    data = extract_software(html_path)

    total = sum(len(items) for items in data.values())
    print(f"\nTotal items extracted: {total}")

    save_to_excel(data, output_path)
    print("Done!")


if __name__ == '__main__':
    main()
