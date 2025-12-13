#!/usr/bin/env python3
"""Extract publication data from existing HTML into Excel spreadsheet.

This is a one-time script to migrate existing HTML data to the spreadsheet format.
"""
import re
from pathlib import Path
from bs4 import BeautifulSoup
import openpyxl


def extract_publications(html_path: Path) -> dict:
    """Extract all publication data from HTML file.

    Returns dict with keys: papers, chapters, dissertations, talks, courses, posters
    """
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f.read(), 'html.parser')

    sections = {
        'papers': 'papers',
        'chapters': 'chapters',
        'dissertations': 'dissertations',
        'talks': 'talks',
        'courses': 'course-mats',
        'posters': 'posters'
    }

    data = {}

    for key, section_id in sections.items():
        section = soup.find('section', id=section_id)
        if not section:
            print(f"Warning: Section '{section_id}' not found")
            data[key] = []
            continue

        cards = section.find_all('div', class_='publication-card')
        items = []

        for card in cards:
            item = extract_card_data(card, key)
            items.append(item)

        data[key] = items
        print(f"Extracted {len(items)} items from '{key}'")

    return data


def extract_card_data(card, section_type: str) -> dict:
    """Extract data from a single publication card."""
    item = {}

    # Get image
    img = card.find('img')
    if img:
        item['image'] = img.get('src', '').replace('images/publications/', '')
    else:
        item['image'] = ''

    # Get title and title URL
    h4 = card.find('h4')
    if h4:
        link = h4.find('a')
        if link:
            item['title'] = link.get_text(strip=True)
            item['title_url'] = link.get('href', '')
        else:
            item['title'] = h4.get_text(strip=True)
            item['title_url'] = ''
    else:
        item['title'] = ''
        item['title_url'] = ''

    # Get citation/description (first <p> that's not publication-links)
    div = card.find('div')
    if div:
        paragraphs = div.find_all('p', recursive=False)
        citation_p = None
        for p in paragraphs:
            if 'publication-links' not in p.get('class', []):
                citation_p = p
                break

        if citation_p:
            # Get inner HTML to preserve formatting
            item['citation'] = get_inner_html(citation_p)
        else:
            item['citation'] = ''
    else:
        item['citation'] = ''

    # Get publication links (PDF, CODE, DATA, etc.)
    links_p = card.find('p', class_='publication-links')
    if links_p:
        item['links_html'] = get_inner_html(links_p)
    else:
        item['links_html'] = ''

    return item


def get_inner_html(element) -> str:
    """Get the inner HTML of an element as a string."""
    return ''.join(str(child) for child in element.children).strip()


def save_to_excel(data: dict, output_path: Path):
    """Save extracted data to Excel spreadsheet with multiple sheets."""
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    sheet_configs = [
        ('papers', ['image', 'title', 'title_url', 'citation', 'links_html']),
        ('chapters', ['image', 'title', 'title_url', 'citation', 'links_html']),
        ('dissertations', ['image', 'title', 'title_url', 'citation', 'links_html']),
        ('talks', ['image', 'title', 'title_url', 'citation', 'links_html']),
        ('courses', ['image', 'title', 'title_url', 'citation', 'links_html']),
        ('posters', ['image', 'title', 'title_url', 'citation', 'links_html']),
    ]

    for sheet_name, columns in sheet_configs:
        ws = wb.create_sheet(title=sheet_name)

        # Write headers
        for col, header in enumerate(columns, 1):
            ws.cell(row=1, column=col, value=header)

        # Write data
        items = data.get(sheet_name, [])
        for row_num, item in enumerate(items, 2):
            for col, header in enumerate(columns, 1):
                value = item.get(header, '')
                ws.cell(row=row_num, column=col, value=value)

        # Adjust column widths
        for col in range(1, len(columns) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 40

    wb.save(output_path)
    print(f"Saved to {output_path}")


def main():
    project_root = Path(__file__).parent.parent
    html_path = project_root / 'publications.html'
    output_path = project_root / 'data' / 'publications.xlsx'

    print(f"Extracting from: {html_path}")
    data = extract_publications(html_path)

    total = sum(len(items) for items in data.values())
    print(f"\nTotal items extracted: {total}")

    save_to_excel(data, output_path)
    print("Done!")


if __name__ == '__main__':
    main()
