#!/usr/bin/env python3
"""Offboard lab members by moving them from active to alumni.

This script:
1. Searches for member(s) by name in people.xlsx
2. Prompts for selection if multiple matches found
3. Moves member from 'members' sheet to 'alumni_undergrads' sheet
4. Updates JRM_CV.tex to add end date to the member's entry
5. Rebuilds people.html

Idempotent: Running twice with same name will detect already-offboarded member.

Usage:
    python offboard_member.py "member name"
    python offboard_member.py "member name" --end-year 2025
    python offboard_member.py --list-no-photo  # List undergrads without photos
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

import openpyxl


def get_project_root() -> Path:
    return Path(__file__).parent.parent


def load_members(xlsx_path: Path) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb["members"]

    headers = [cell.value for cell in sheet[1]]

    members = []
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not any(row):
            continue
        member: Dict[str, Any] = {"_row_idx": row_idx}
        for header, value in zip(headers, row):
            if header is not None:
                member[str(header)] = value if value is not None else ""
        members.append(member)

    wb.close()
    return members


def load_alumni_undergrads(xlsx_path: Path) -> List[str]:
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb["alumni_undergrads"]

    alumni_names = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            alumni_names.append(str(row[0]).lower())

    wb.close()
    return alumni_names


def member_already_alumni(xlsx_path: Path, name: str) -> bool:
    alumni_names = load_alumni_undergrads(xlsx_path)
    return name.lower() in alumni_names or name.title().lower() in alumni_names


def find_members_by_name(
    members: List[Dict[str, Any]], search_name: str
) -> List[Dict[str, Any]]:
    search_lower = search_name.lower()
    return [m for m in members if search_lower in m.get("name", "").lower()]


def prompt_for_selection(matches: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    print(f"\nFound {len(matches)} matching members:")
    for i, m in enumerate(matches, 1):
        role = m.get("role", "unknown role")
        image = m.get("image", "")
        has_photo = "has photo" if image else "NO PHOTO"
        print(f"  {i}. {m['name']} ({role}) [{has_photo}]")

    print(f"  0. Cancel")

    while True:
        try:
            choice = input("\nSelect member number (or 0 to cancel): ").strip()
            if not choice:
                continue
            choice_num = int(choice)
            if choice_num == 0:
                return None
            if 1 <= choice_num <= len(matches):
                return matches[choice_num - 1]
            print(f"Please enter a number between 0 and {len(matches)}")
        except ValueError:
            print("Please enter a valid number")


def move_to_alumni(xlsx_path: Path, member: Dict[str, Any], years_string: str) -> bool:
    """Move member to alumni sheet with the given years string (e.g., '2024-2026' or '2026')."""
    if member_already_alumni(xlsx_path, member["name"]):
        print(
            f"  {member['name']} is already in alumni list, skipping spreadsheet update"
        )
        return True

    wb = openpyxl.load_workbook(xlsx_path)

    members_sheet = wb["members"]
    members_sheet.delete_rows(member["_row_idx"])

    alumni_sheet = wb["alumni_undergrads"]

    alumni_sheet.insert_rows(2)
    alumni_sheet.cell(row=2, column=1, value=member["name"].title())
    alumni_sheet.cell(row=2, column=2, value=years_string)

    wb.save(xlsx_path)
    wb.close()
    print(f"  Moved {member['name']} to alumni_undergrads with years {years_string}")
    return True


def get_cv_start_year(cv_path: Path, member_name: str) -> Optional[str]:
    """Extract the start year from the CV entry for a member."""
    content = cv_path.read_text(encoding="utf-8")
    name_escaped = re.escape(member_name)

    # Match open entry: \item Name[*]? (start_year -- )
    pattern_open = r"\\item\s+" + name_escaped + r"\*?\s*\((\d{4})\s*--\s*\)"
    match = re.search(pattern_open, content, re.IGNORECASE)
    if match:
        return match.group(1)

    # Match closed entry: \item Name[*]? (start_year -- end_year) or (year)
    pattern_closed = r"\\item\s+" + name_escaped + r"\*?\s*\((\d{4})(?:\s*--\s*\d{4})?\)"
    match = re.search(pattern_closed, content, re.IGNORECASE)
    if match:
        return match.group(1)

    return None


def cv_entry_already_closed(cv_path: Path, member_name: str) -> bool:
    content = cv_path.read_text(encoding="utf-8")
    name_escaped = re.escape(member_name)
    # Match entries that already have an end year: (YYYY) or (YYYY -- YYYY)
    pattern_closed = r"\\item\s+" + name_escaped + r"\*?\s*\(\d{4}(\s*--\s*\d{4})?\)"
    pattern_open = r"\\item\s+" + name_escaped + r"\*?\s*\(\d{4}\s*--\s*\)"

    has_closed = bool(re.search(pattern_closed, content, re.IGNORECASE))
    has_open = bool(re.search(pattern_open, content, re.IGNORECASE))

    return has_closed and not has_open


def update_cv_entry(cv_path: Path, member_name: str, end_year: str) -> bool:
    if cv_entry_already_closed(cv_path, member_name):
        print(f"  CV entry for {member_name} already has end date, skipping")
        return True

    content = cv_path.read_text(encoding="utf-8")

    name_escaped = re.escape(member_name)

    # Regex: \item Name[*]? (start_year -- )
    pattern = r"(\\item\s+)(" + name_escaped + r")(\*?)(\s*\((\d{4})\s*--\s*\))"

    def replace_entry(match):
        prefix = match.group(1)
        name = match.group(2)
        asterisk = match.group(3)
        start_year = match.group(5)

        if start_year == end_year:
            return f"{prefix}{name}{asterisk} ({end_year})"
        else:
            return f"{prefix}{name}{asterisk} ({start_year} -- {end_year})"

    new_content, count = re.subn(pattern, replace_entry, content, flags=re.IGNORECASE)

    if count > 0:
        cv_path.write_text(new_content, encoding="utf-8")
        print(f"  Updated CV entry for {member_name}")
        return True
    else:
        print(f"  WARNING: Could not find open CV entry for {member_name}")
        return False


def list_undergrads_without_photos(xlsx_path: Path) -> None:
    members = load_members(xlsx_path)

    no_photo = [
        m
        for m in members
        if m.get("role", "").lower() == "undergrad" and not m.get("image")
    ]

    if not no_photo:
        print("All undergraduate members have photos!")
        return

    print(f"\nUndergraduate members without photos ({len(no_photo)}):")
    for m in no_photo:
        print(f"  - {m['name']}")


def offboard_member(
    search_name: str, end_year: Optional[str] = None, skip_confirm: bool = False
) -> bool:
    project_root = get_project_root()
    xlsx_path = project_root / "data" / "people.xlsx"
    cv_path = project_root / "documents" / "JRM_CV.tex"

    if end_year is None:
        end_year = str(datetime.now().year)

    if member_already_alumni(xlsx_path, search_name):
        print(
            f"'{search_name}' is already in the alumni list (idempotent - no changes needed)"
        )
        return True

    members = load_members(xlsx_path)
    matches = find_members_by_name(members, search_name)

    if not matches:
        print(f"No active members found matching '{search_name}'")
        return False

    if len(matches) == 1:
        member = matches[0]
        print(f"\nFound: {member['name']} ({member.get('role', 'unknown role')})")
    else:
        member = prompt_for_selection(matches)
        if member is None:
            print("Cancelled")
            return False

    if not skip_confirm:
        confirm = (
            input(f"\nOffboard {member['name']} with end year {end_year}? [y/N]: ")
            .strip()
            .lower()
        )
        if confirm != "y":
            print("Cancelled")
            return False

    print(f"\nOffboarding {member['name']}...")

    # Get start year from CV to create the full date range for alumni sheet
    start_year = get_cv_start_year(cv_path, member["name"])
    if start_year and start_year != end_year:
        years_string = f"{start_year}-{end_year}"
    else:
        years_string = end_year

    move_to_alumni(xlsx_path, member, years_string)
    update_cv_entry(cv_path, member["name"], end_year)

    # Update lab-manual (best-effort; failure doesn't block offboarding)
    try:
        from parse_lab_manual import move_member_to_alumni as lm_move, commit_and_push_lab_manual
        lab_manual_tex = project_root / 'lab-manual' / 'lab_manual.tex'
        if lab_manual_tex.exists():
            print("  Updating lab-manual...")
            lm_move(lab_manual_tex, member["name"], end_year)
            try:
                commit_and_push_lab_manual(
                    project_root / 'lab-manual',
                    f"Offboard {member['name']}"
                )
                print(f"  Updated lab-manual and pushed to remote")
            except RuntimeError as e:
                print(f"  WARNING: Lab-manual updated locally but push failed: {e}")
        else:
            print("  NOTE: Lab-manual submodule not found, skipping lab-manual update")
    except Exception as e:
        print(f"  WARNING: Could not update lab-manual: {e}")

    print(f"\nSuccessfully offboarded {member['name']}")
    print("Run 'python build.py' to rebuild people.html")

    return True


def main():
    parser = argparse.ArgumentParser(
        description="Offboard lab members by moving them from active to alumni.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python offboard_member.py "john doe"
    python offboard_member.py "john" --end-year 2024
    python offboard_member.py --list-no-photo
    python offboard_member.py "jane" -y  # Skip confirmation
        """,
    )

    parser.add_argument(
        "name", nargs="?", help="Name (or partial name) of member to offboard"
    )
    parser.add_argument(
        "--end-year",
        default=str(datetime.now().year),
        help=f"End year for alumni entry (default: {datetime.now().year})",
    )
    parser.add_argument(
        "-y", "--yes", action="store_true", help="Skip confirmation prompt"
    )
    parser.add_argument(
        "--list-no-photo",
        action="store_true",
        help="List all undergraduate members without photos",
    )

    args = parser.parse_args()

    project_root = get_project_root()
    xlsx_path = project_root / "data" / "people.xlsx"

    if args.list_no_photo:
        list_undergrads_without_photos(xlsx_path)
        return

    if not args.name:
        parser.print_help()
        sys.exit(1)

    success = offboard_member(args.name, args.end_year, args.yes)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
